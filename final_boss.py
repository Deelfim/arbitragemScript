from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import pandas as pd
import time
import os
import json
import html 
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import Counter
from datetime import datetime
import winreg

# ==========================================
# CONFIGURAÇÕES DE ACESSO E DADOS
# ==========================================
NOME_UTILIZADOR = "gabrielcorreiadelfim@gmail.com"  
PASSWORD = "#Areiased5"      
URL_LOGIN = "https://juizesbasquetebol.pt/login"  
URL_JOGOS = "https://juizesbasquetebol.pt/dashboard/os-meus-jogos/jogos-realizados" 

MEU_NOME_SISTEMA = "Delfim Gabriel Almeida Correia"
MEU_NOME_CURTO = "Delfim" 

# Dados para o Mapa de Contas
MEU_IBAN = "PT50 0000 0000 0000 0000 0" 
LOCAL_PARTIDA = "Vale de Cambra"

PASTA_DESTINO = r"C:\Users\Gabriel\Documents\vscode\scrpit"
FICHEIRO_FINAL = os.path.join(PASTA_DESTINO, "Dashboard_Arbitragem.xlsx")

# --- OS 3 FICHEIROS DE SUPORTE ---
CAMINHO_PAVILHOES = os.path.join(PASTA_DESTINO, "pavilhoes.xlsx")
CAMINHO_TABELA_KMS = os.path.join(PASTA_DESTINO, "Tabela_KMs.xlsx")
CAMINHO_TABELA_EUROS = os.path.join(PASTA_DESTINO, "Tabela_Euros.xlsx")

# ==========================================
# TABELAS DE PREÇOS (PDF 25/26)
# ==========================================
TABELA_PREMIOS_OFICIAL = {
    "Oficial de Mesa": {
        "Nacional": {
            "Masculino": {
                "LPB": 39.0, "Proliga": 25.0, "CN1": 24.0, "CN2": 23.0, 
                "Sub18": 19.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0
            },
            "Feminino": {
                "Liga Fem": 18.0, "CN1": 16.0, "Sub18": 12.0, "Sub16": 12.0, "Sub14": 12.0, "Outros": 0
            }
        },
        "Associação": {
            "Masculino": {"Seniores": 19.0, "Masters": 14.0, "Sub21": 12.0, "Sub18": 10.0, "Sub16": 9.0, "Sub14": 8.0, "Outros": 0},
            "Feminino": {"Seniores": 19.0, "Masters": 14.0, "Sub21": 12.0, "Sub18": 10.0, "Sub16": 9.0, "Sub14": 8.0, "Outros": 0}
        }
    }
}

# ==========================================
# MOTOR DE CÁLCULO E CLASSIFICAÇÃO
# ==========================================

# 1. Carregar Pavilhões do Novo Excel
def carregar_mapa_pavilhoes():
    mapa = {}
    if os.path.exists(CAMINHO_PAVILHOES):
        try:
            df_p = pd.read_excel(CAMINHO_PAVILHOES)
            for _, row in df_p.iterrows():
                nome_pav = str(row['PAVILHÃO']).strip().lower()
                localidade = str(row['LOCALIDADE']).strip()
                mapa[nome_pav] = localidade
        except Exception as e:
            print(f"Erro ao ler pavilhões: {e}")
    return mapa

MAPA_EXTERNO_PAVILHOES = carregar_mapa_pavilhoes()

# 2. Obter Localidade Real a partir do Mapa
def obter_localidade_real(pavilhao_portal):
    pav_limpo = str(pavilhao_portal).lower().strip()
    for nome_excel, localidade in MAPA_EXTERNO_PAVILHOES.items():
        if nome_excel in pav_limpo or pav_limpo in nome_excel: 
            return localidade
    return pavilhao_portal

def classificar_jogo(row):
    comp = str(row.get('Competição', '')).lower()
    tipo = 'Associação'
    if 'master' in comp: tipo = 'Masters'
    elif any(x in comp for x in ["nacional", "festa", "lpb", "proliga", "betclic"]): tipo = 'Nacional'
    
    gen = "Feminino" if "fem" in comp else "Masculino"
    
    esc = "Outros"
    if "sub 18" in comp or "sub18" in comp: esc = "Sub18"
    elif "sub 16" in comp or "sub16" in comp: esc = "Sub16"
    elif "sub 14" in comp or "sub14" in comp: esc = "Sub14"
    elif "master" in comp: esc = "Masters"
    elif any(x in comp for x in ["liga betclic", "lpb"]): esc = "LPB"
    elif "proliga" in comp: esc = "Proliga"
    elif "1ª div" in comp: esc = "CN1"
    elif "2ª div" in comp: esc = "CN2"
    elif "senior" in comp or "sénior" in comp: esc = "Seniores"
    
    return tipo, gen, esc

def calcular_financeiro(df):
    if df.empty: return df
    df = df.copy() 
    
    # Prevenção de Erros nas Colunas
    df = df.loc[:, ~df.columns.duplicated()]
    colunas_a_apagar = ['Tipo', 'Género', 'Escalão_Real', 'Data_dt', 'Localidade_Destino', 'Mês', 'KMs', 'Deslocação (€)', 'Prémio (€)', 'Total (€)']
    df.drop(columns=[c for c in colunas_a_apagar if c in df.columns], inplace=True, errors='ignore')
    
    lista_tipos, lista_generos, lista_escaloes = [], [], []
    for _, row in df.iterrows():
        t, g, e = classificar_jogo(row)
        lista_tipos.append(t); lista_generos.append(g); lista_escaloes.append(e)
        
    df['Tipo'] = lista_tipos
    df['Género'] = lista_generos
    df['Escalão_Real'] = lista_escaloes

    # Cálculo dos Prémios
    premios = []
    for _, row in df.iterrows():
        tipo_b = "Associação" if row['Tipo'] == "Masters" else row['Tipo']
        try:
            val = TABELA_PREMIOS_OFICIAL["Oficial de Mesa"][tipo_b][row['Género']].get(row['Escalão_Real'], 0.0)
            if val == 0: val = TABELA_PREMIOS_OFICIAL["Oficial de Mesa"][tipo_b]["Masculino"].get(row['Escalão_Real'], 0.0)
        except: val = 0.0
        premios.append(val)
    df['Prémio (€)'] = premios
    
    # Identifica o Destino
    df['Localidade_Destino'] = df['Recinto'].apply(obter_localidade_real)
    df['KMs'] = 0.0
    df['Deslocação (€)'] = 0.0
    
    df_e = pd.DataFrame()
    df_k = pd.DataFrame()
    
    # Carrega os ficheiros de KMs e Euros
    if os.path.exists(CAMINHO_TABELA_EUROS):
        try:
            df_e = pd.read_excel(CAMINHO_TABELA_EUROS, index_col=0)
            df_e.index, df_e.columns = df_e.index.astype(str).str.strip(), df_e.columns.astype(str).str.strip()
        except: pass

    if os.path.exists(CAMINHO_TABELA_KMS):
        try:
            df_k = pd.read_excel(CAMINHO_TABELA_KMS, index_col=0)
            df_k.index, df_k.columns = df_k.index.astype(str).str.strip(), df_k.columns.astype(str).str.strip()
        except: pass

    # Procura os valores de Deslocação
    for idx, row in df.iterrows():
        origem = LOCAL_PARTIDA.strip()
        destino = str(row['Localidade_Destino']).strip()
        
        # KMs
        if not df_k.empty and destino in df_k.columns and origem in df_k.index:
            try: df.at[idx, 'KMs'] = float(str(df_k.at[origem, destino]).replace(',','.'))
            except: pass
            
        # Euros
        if not df_e.empty and destino in df_e.columns and origem in df_e.index:
            try: df.at[idx, 'Deslocação (€)'] = float(str(df_e.at[origem, destino]).replace('€','').replace(',','.').strip())
            except: pass

    # Regra de não pagar viagens repetidas no mesmo dia/local
    df['Data_dt'] = pd.to_datetime(df['Data'], format="%d-%m-%Y")
    for (data, local), grupo in df.groupby(['Data_dt', 'Localidade_Destino']):
        paga_idx = grupo[grupo['Tipo'] == "Nacional"].index[0] if "Nacional" in grupo['Tipo'].values else grupo.index[0]
        for idx in grupo.index:
            if idx != paga_idx: 
                df.at[idx, 'Deslocação (€)'] = 0.0
                df.at[idx, 'KMs'] = 0.0
            
    df['Total (€)'] = df['Prémio (€)'] + df['Deslocação (€)']
    df['Mês'] = df['Data_dt'].dt.strftime('%m-%Y')
    return df

# ==========================================
# FUNÇÕES DE FORMATAÇÃO PREMIUM (OPENPYXL)
# ==========================================
def formatar_aba_geral(ws):
    azul_escuro = PatternFill(start_color="1F4E78", fill_type="solid")
    azul_header = PatternFill(start_color="2F75B5", fill_type="solid")
    zebra = PatternFill(start_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))
    
    ws.insert_rows(1, 4)
    ws.merge_cells('A1:I2')
    ws['A1'] = "BASE DE DADOS GLOBAL - JOGOS PORTAL"
    ws['A1'].font = Font(name='Segoe UI', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = azul_escuro
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = azul_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r in range(6, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if r % 2 == 0: cell.fill = zebra
            if c in [1, 2, 3, 9]: cell.alignment = Alignment(horizontal="center")

    ws.auto_filter.ref = f"A5:{openpyxl.utils.get_column_letter(ws.max_column)}5"
    ws.freeze_panes = "A6"
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

def formatar_mapa_contas(ws, titulo):
    font_bold = Font(name='Segoe UI', bold=True)
    fill_cap = PatternFill(start_color="B4C6E7", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    max_col = ws.max_column
    letra_max_col = openpyxl.utils.get_column_letter(max_col)
    
    ws.insert_rows(1, 8)
    ws.merge_cells(f'A1:{letra_max_col}1'); ws['A1'] = titulo; ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells(f'A2:{letra_max_col}2'); ws['A2'] = "MAPA DE PRESTAÇÃO DE CONTAS - SERVIÇOS / DESLOCAÇÕES"; ws['A2'].font = font_bold
    ws['A4'] = f"Nome: {MEU_NOME_SISTEMA}"; ws['A5'] = f"Localidade: {LOCAL_PARTIDA}"; ws['A6'] = f"IBAN: {MEU_IBAN}"
    ws['F4'] = "Categoria: Oficial de Mesa / Árbitro"; ws['F4'].font = font_bold
    
    for cell in ws[9]:
        cell.font = font_bold; cell.fill = fill_cap; cell.alignment = Alignment(horizontal="center"); cell.border = border
        
    for r in range(10, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c); cell.border = border
            if r % 2 == 0: cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
            if c in [9, 10, 11]: cell.number_format = '#,##0.00 €'

    max_r = ws.max_row
    ws.auto_filter.ref = f"A9:{letra_max_col}9"
    ws.freeze_panes = "A10"
    
    linha_totais = max_r + 2
    ws[f'J{linha_totais}'] = "A RECEBER:"; ws[f'J{linha_totais}'].font = font_bold
    ws[f'K{linha_totais}'] = f"=SUBTOTAL(109, K10:K{max_r})"; ws[f'K{linha_totais}'].number_format = '#,##0.00 €'; ws[f'K{linha_totais}'].font = font_bold
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

def formatar_estatisticas_gerais(ws, is_resumo=False):
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F2F2", fill_type="solid")
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.font = font_header; cell.fill = fill_header; cell.alignment = Alignment(horizontal="center")
            else:
                if cell.row % 2 == 0: cell.fill = fill_zebra
                
            if is_resumo:
                cabecalho = str(ws.cell(row=1, column=cell.column).value)
                if "Total" in cabecalho or "Prémio" in cabecalho or "Deslocação" in cabecalho:
                    if cell.row > 1 and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00 €'
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 3, 40)

# ==========================================
# MOTOR PRINCIPAL
# ==========================================
def extrair_e_gerar():
    if os.path.exists(FICHEIRO_FINAL):
        try:
            with open(FICHEIRO_FINAL, 'a'): pass
        except PermissionError:
            print("\n" + "="*60)
            print("❌ ERRO: O TEU FICHEIRO EXCEL ESTÁ ABERTO! ❌")
            print(f"O ficheiro '{FICHEIRO_FINAL}' está aberto no Excel.")
            print("Por favor, fecha o Excel completamente e volta a correr o código.")
            print("="*60 + "\n")
            return

    print("A iniciar o robô do Chrome...")
    chrome_options = Options()
    
    # --- CÓDIGO UNIVERSAL PARA ACHAR O CHROME OFICIAL EM QUALQUER PC ---
    caminho_chrome = ""
    try:
        # Tenta procurar no registo do Windows onde o Chrome foi instalado
        chave = winreg.OpenKey(winreg.HKEY_LOCAL_MECHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe")
        caminho_chrome = winreg.QueryValue(chave, None)
    except:
        # Se não encontrar no registo, tenta os caminhos padrão
        caminhos_padrao = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
        ]
        for p in caminhos_padrao:
            if os.path.exists(p):
                caminho_chrome = p
                break
                
    if caminho_chrome:
        print(f"-> A usar o teu Chrome Oficial para evitar bloqueios do Windows.")
        chrome_options.binary_location = caminho_chrome
    else:
        print("-> Chrome Oficial não encontrado. A tentar a versão invisível...")

    chrome_options.add_argument("--no-sandbox") 
    chrome_options.add_argument("--disable-dev-shm-usage") 
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"]) 
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    try:
        driver = webdriver.Chrome(options=chrome_options)
        wait = WebDriverWait(driver, 15)
    except Exception as e:
        print(f"\n❌ Erro crítico ao abrir o Chrome: {e}")
        print("Tenta fechar TODAS as janelas abertas do Google Chrome antes de correr o código.")
        return

    todos_os_jogos = []
    
    jogos_conhecidos = set()
    df_antigo = pd.DataFrame()
    if os.path.exists(FICHEIRO_FINAL):
        try:
            df_teste = pd.read_excel(FICHEIRO_FINAL, sheet_name="Geral", nrows=5)
            if "BASE" in str(df_teste.columns[0]):
                df_antigo = pd.read_excel(FICHEIRO_FINAL, sheet_name="Geral", skiprows=4)
            else:
                df_antigo = pd.read_excel(FICHEIRO_FINAL, sheet_name="Geral")
            
            if 'Nº Jogo' in df_antigo.columns:
                jogos_conhecidos = set(df_antigo['Nº Jogo'].astype(str).tolist())
                print(f"-> Detetados {len(jogos_conhecidos)} jogos antigos.")
        except: pass

    try:
        driver.get(URL_LOGIN)
        wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(NOME_UTILIZADOR)
        driver.find_element(By.ID, "password").send_keys(PASSWORD)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(5); driver.get(URL_JOGOS); time.sleep(4)
        
        pagina = 1
        ja_existe = False
        while True:
            print(f"A extrair Página {pagina}...")
            wait.until(lambda d: d.execute_script("return typeof Fabrik !== 'undefined';"))
            time.sleep(1)
            
            dados = driver.execute_script("return Fabrik.blocks['list_5_com_fabrik_5'].options.data;")
            if not dados or len(dados[0]) == 0: break
            
            for j in dados[0]:
                row = j['data']
                n_jogo = str(row.get('fab_jogos___num_jogo_raw', ''))
                
                if n_jogo in jogos_conhecidos:
                    print(f"\n>>> Jogo {n_jogo} já lido. Sincronização concluída!")
                    ja_existe = True
                    break
                
                comp_limpa = html.unescape(row.get('fab_jogos___competicao', ''))
                fase_limpa = html.unescape(row.get('fab_jogos___fase', ''))
                rec_limpo = html.unescape(row.get('fab_jogos___recinto', ''))
                
                juizes = json.loads(row.get('fab_nomeacoes___juiz', '[]'))
                funcoes = json.loads(row.get('fab_nomeacoes___funcao_jogo', '[]'))
                eq_dict = {f: "" for f in ["Árbitro Principal", "Árbitro Auxiliar 1", "Árbitro Auxiliar 2", "Marcador", "Marcador Auxiliar", "Cronometrista", "Operador 24\""]}
                minha_f = "N/A"
                
                for n, f in zip(juizes, funcoes):
                    n_l = html.unescape(n)
                    if MEU_NOME_CURTO.lower() in n_l.lower(): minha_f = f
                    if f in eq_dict: eq_dict[f] = n_l
                
                todos_os_jogos.append({
                    "Data": row.get('fab_jogos___data', ''), "Hora": row.get('fab_jogos___hora', ''),
                    "Nº Jogo": n_jogo, "Recinto": rec_limpo,
                    "Equipas": f"{html.unescape(row.get('fab_jogos___visitado',''))} vs {html.unescape(row.get('fab_jogos___visitante',''))}",
                    "Competição": comp_limpa, "Fase": fase_limpa,
                    "Minha Função": minha_f, **eq_dict
                })
            
            if ja_existe: break
                
            try:
                btn = driver.find_element(By.XPATH, "//a[@rel='next' or contains(text(), '»')]")
                url = btn.get_attribute("href")
                if not url or url.endswith("#"): break
                driver.get(url); time.sleep(4); pagina += 1
            except: break

        if todos_os_jogos or not df_antigo.empty:
            print("\nA organizar pastas e calcular valores...")
            if todos_os_jogos:
                df_novos = pd.DataFrame(todos_os_jogos)
                df_completo = pd.concat([df_novos, df_antigo], ignore_index=True)
                df_completo = df_completo.drop_duplicates(subset=['Nº Jogo'], keep='first')
            else:
                df_completo = df_antigo.copy()

            df_completo = calcular_financeiro(df_completo)
            
            df_resumo = df_completo.groupby('Mês')[['Prémio (€)', 'Deslocação (€)', 'Total (€)']].sum().reset_index()
            df_resumo['Mês_dt'] = pd.to_datetime(df_resumo['Mês'], format='%m-%Y')
            df_resumo = df_resumo.sort_values('Mês_dt').drop(columns=['Mês_dt'])
            
            col_mapa = ["Mês", "Data", "Competição", "Escalão_Real", "Minha Função", "Recinto", "Equipas", "KMs", "Deslocação (€)", "Prémio (€)", "Total (€)"]
            df_nac = df_completo[df_completo['Tipo'] == "Nacional"][col_mapa].rename(columns={"Competição":"Prova","Escalão_Real":"Escalão","Equipas":"Jogo"})
            df_aba = df_completo[df_completo['Tipo'] == "Associação"][col_mapa].rename(columns={"Competição":"Prova","Escalão_Real":"Escalão","Equipas":"Jogo"})
            df_mas = df_completo[df_completo['Tipo'] == "Masters"][col_mapa].rename(columns={"Competição":"Prova","Escalão_Real":"Escalão","Equipas":"Jogo"})
            df_geral = df_completo.drop(columns=['Tipo', 'Género', 'Escalão_Real', 'Data_dt', 'Localidade_Destino', 'Mês'], errors='ignore')

            stats_funcao = df_completo['Minha Função'].value_counts().reset_index(); stats_funcao.columns = ['Função', 'Total Jogos']
            stats_recinto = df_completo['Recinto'].value_counts().reset_index(); stats_recinto.columns = ['Pavilhão', 'Total Jogos']
            stats_comp = df_completo['Competição'].value_counts().reset_index(); stats_comp.columns = ['Competição', 'Total Jogos']
            
            colegas = []
            for col in ["Árbitro Principal", "Árbitro Auxiliar 1", "Árbitro Auxiliar 2", "Marcador", "Marcador Auxiliar", "Cronometrista", "Operador 24\""]:
                for val in df_completo[col].dropna():
                    if val and val.strip() and MEU_NOME_CURTO.lower() not in val.lower(): colegas.append(val.strip())
            df_equipa = pd.DataFrame(Counter(colegas).most_common(), columns=['Colega', 'Nº de Jogos'])

            with pd.ExcelWriter(FICHEIRO_FINAL, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name="Resumo Financeiro", index=False)
                df_geral.to_excel(writer, sheet_name="Geral", index=False)
                df_nac.to_excel(writer, sheet_name="Nacional", index=False)
                df_aba.to_excel(writer, sheet_name="Associação", index=False)
                if not df_mas.empty: df_mas.to_excel(writer, sheet_name="Masters", index=False)
                
                stats_funcao.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=0)
                stats_recinto.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=3)
                stats_comp.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=6)
                
                df_equipa.to_excel(writer, sheet_name="Minha Equipa", index=False)

            wb = openpyxl.load_workbook(FICHEIRO_FINAL)
            
            if "Resumo Financeiro" in wb.sheetnames: formatar_estatisticas_gerais(wb["Resumo Financeiro"], is_resumo=True)
            formatar_aba_geral(wb["Geral"])
            formatar_mapa_contas(wb["Nacional"], "FEDERAÇÃO PORTUGUESA DE BASQUETEBOL")
            formatar_mapa_contas(wb["Associação"], "ASSOCIAÇÃO DE BASQUETEBOL DE AVEIRO")
            if "Masters" in wb.sheetnames: formatar_mapa_contas(wb["Masters"], "ASSOCIAÇÃO DE AVEIRO - JOGOS MASTERS")
            
            for ws_name in ["Estatísticas", "Minha Equipa"]:
                if ws_name in wb.sheetnames: formatar_estatisticas_gerais(wb[ws_name], is_resumo=False)
            
            wb.save(FICHEIRO_FINAL)
            print(f"\n✅ O Excel finalizado e pronto!")
            
    except Exception as e: print(f"Erro fatal: {e}")
    finally: 
        try: driver.quit()
        except: pass

if __name__ == "__main__": extrair_e_gerar()