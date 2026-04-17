import requests
import pandas as pd
from io import StringIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# ==========================================
# CONFIGURACOES PRINCIPAIS
# ==========================================
MES_PROCESSAMENTO = 1
ANO_PROCESSAMENTO = 2026
NOME_COMPLETO = "Delfim Correia"  # PREENCHE AQUI
MEU_IBAN = "PT50 0000 0000 0000 0000 0"    # PREENCHE AQUI
LOCAL_PARTIDA = "Vale de Cambra"
MINHA_FUNCAO = "Oficial de Mesa" 
FASE_ATUAL = "1FASE" # "1FASE" ou "2FASE"
PASTA_DESTINO = r"C:\Users\Gabriel\Documents\vscode\scrpit"
URL_FPB = "https://www.fpb.pt/juizes/18534/"

CAMINHO_TABELA_KMS = os.path.join(PASTA_DESTINO, "Tabela_KMs.xlsx")
CAMINHO_TABELA_EUROS = os.path.join(PASTA_DESTINO, "Tabela_Euros.xlsx")
CAMINHO_TEMPLATE_NACIONAL = os.path.join(PASTA_DESTINO, "deslocaçoes_nacional.xlsx")
CAMINHO_TEMPLATE_ABA = os.path.join(PASTA_DESTINO, "despesas_distrital.xlsx")

# ==========================================
# 1. TABELAS DE PREMIOS E MAPAS
# ==========================================
TABELA_PREMIOS = {
    "Oficial de Mesa": {
        "Nacional": {
            "Masculino": {
                "1ª Fase": { "CN1": 25.0, "CN2": 24.0, "Sub18": 19.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0 },
                "2ª Fase": { "CN1": 26.0, "CN2": 25.0, "Sub18": 19.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0 }
            },
            "Feminino": {
                "1ª Fase": { "CN1": 19.0, "CN2": 17.0, "Sub18": 17.0, "Sub16": 15.0, "Sub14": 15.0, "Outros": 0 },
                "2ª Fase": { "CN1": 20.0, "CN2": 18.0, "Sub18": 17.0, "Sub16": 15.0, "Sub14": 15.0, "Outros": 0 }
            }
        },
        "Associação": {
            "Masculino": { "Seniores": 19.0, "Masters": 19.0, "Sub21": 12.0, "Sub18": 10.0, "Sub16": 9.0, "Sub14": 8.0, "Outros": 0 },
            "Feminino": { "Seniores": 19.0, "Masters": 19.0, "Sub21": 12.0, "Sub18": 10.0, "Sub16": 9.0, "Sub14": 8.0, "Outros": 0 }
        }
    },
    "Arbitro": {
        "Nacional": {
            "Masculino": {
                "1ª Fase": { "CN1": 46.0, "CN2": 35.0, "Sub18": 28.0, "Sub16": 25.0, "Sub14": 25.0, "Outros": 0 },
                "2ª Fase": { "CN1": 47.0, "CN2": 38.0, "Sub18": 26.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0 }
            },
            "Feminino": {
                "1ª Fase": { "CN1": 34.0, "CN2": 24.0, "Sub18": 23.0, "Sub16": 20.0, "Sub14": 18.0, "Outros": 0 },
                "2ª Fase": { "CN1": 36.0, "CN2": 25.0, "Sub18": 23.0, "Sub16": 20.0, "Sub14": 18.0, "Outros": 0 }
            }
        },
        "Associação": {
            "Masculino": { "Seniores": 27.0, "Masters": 27.0, "Sub21": 14.0, "Sub18": 13.0, "Sub16": 11.0, "Sub14": 10.0, "Outros": 0 },
            "Feminino": { "Seniores": 27.0, "Masters": 27.0, "Sub21": 14.0, "Sub18": 13.0, "Sub16": 11.0, "Sub14": 10.0, "Outros": 0 }
        }
    }
}

MAPA_PAVILHOES = {
    "arena de ovar": "Ovar", "ovar": "Ovar",
    "municipal de anadia": "Anadia", "acr cerca": "Anadia", "s. pedro": "Anadia", "anadia": "Anadia",
    "ventosa do bairro": "Mealhada", "mealhada": "Mealhada",
    "vale cambra": "Vale de Cambra", "vale de cambra": "Vale de Cambra",
    "paulo pinto": "S.João da Madeira", "são joão da madeira": "S.João da Madeira", "sao joao da madeira": "S.João da Madeira", "s. joão": "S.João da Madeira", "s. joao": "S.João da Madeira",
    "municipal de vagos": "Vagos", "vagos": "Vagos",
    "oliveira do bairro": "Oliveira do Bairro",
    "galitos": "Aveiro", "joão afonso": "Aveiro", "joao afonso": "Aveiro", "aveiro": "Aveiro",
    "secundária de arouca": "Arouca", "secundaria de arouca": "Arouca", "arouca": "Arouca",
    "albergaria": "Albergaria a Velha",
    "esgueira": "Esgueira", "clube do povo de esgueira": "Esgueira",
    "costeira": "Oliveira de Azeméis", "antónio costeira": "Oliveira de Azeméis", "antonio costeira": "Oliveira de Azeméis", "salvador machado": "Oliveira de Azeméis", "oliveira de azeméis": "Oliveira de Azeméis", "oliveira de azemeis": "Oliveira de Azeméis", "ferreira castro": "Oliveira de Azeméis",
    "águeda": "Águeda", "agueda": "Águeda", "gica": "Águeda",
    "gafanha": "Gafanha da Nazaré",
    "antónio júlio silva": "Paços de Brandão", "antonio julio silva": "Paços de Brandão", "paços de brandão": "Paços de Brandão", "pacos de brandao": "Paços de Brandão",
    "adriano nordeste": "Ílhavo", "illiabum": "Ílhavo", "ílhavo": "Ílhavo", "ilhavo": "Ílhavo",
    "estarreja": "Estarreja",
    "sangalhos": "Sangalhos", "complexo desportivo de sangalhos": "Sangalhos",
    "luso": "Luso", "calvão": "Calvão", "calvao": "Calvão"
}

# ==========================================
# CARREGAR BASES DE DADOS
# ==========================================
DF_KMS = None
DF_EUROS = None
caminho_csv_kms = CAMINHO_TABELA_KMS.replace(".xlsx", ".csv")
caminho_csv_euros = CAMINHO_TABELA_EUROS.replace(".xlsx", ".csv")

print("\n--- A VERIFICAR MATRIZ DE DESLOCACOES ---")
try:
    if os.path.exists(CAMINHO_TABELA_KMS):
        DF_KMS = pd.read_excel(CAMINHO_TABELA_KMS, index_col=0)
        DF_EUROS = pd.read_excel(CAMINHO_TABELA_EUROS, index_col=0)
        print("-> Matrizes carregadas.")
    elif os.path.exists(caminho_csv_kms):
        DF_KMS = pd.read_csv(caminho_csv_kms, index_col=0)
        DF_EUROS = pd.read_csv(caminho_csv_euros, index_col=0)
        print("-> Matrizes carregadas (CSV).")
    else:
        print(f"-> AVISO CRÍTICO: Não encontrei as tabelas em {PASTA_DESTINO}!")
        
    if DF_KMS is not None:
        DF_KMS.index = DF_KMS.index.astype(str).str.strip()
        DF_KMS.columns = DF_KMS.columns.astype(str).str.strip()
        DF_EUROS.index = DF_EUROS.index.astype(str).str.strip()
        DF_EUROS.columns = DF_EUROS.columns.astype(str).str.strip()
except Exception as e:
    print(f"-> AVISO: Erro ao tentar ler as tabelas: {e}")
print("-----------------------------\n")

# ==========================================
# FUNCOES INTELIGENTES
# ==========================================
def definir_tipo_competicao(nome_prova):
    p = str(nome_prova).lower()
    if 'master' in p or 'distrital' in p or 'inter-associa' in p: return 'Associação'
    if 'nacional' in p: return 'Nacional'
    return 'Associação'

def extrair_genero(nome_prova):
    return 'Feminino' if 'fem' in str(nome_prova).lower() else 'Masculino'

def extrair_fase(nome_prova):
    p = str(nome_prova).lower()
    if '2ª fase' in p or '2a fase' in p or 'ii fase' in p or 'final' in p: return '2ª Fase'
    return '1ª Fase'

def extrair_escalao(nome_prova, tipo_jogo):
    p = str(nome_prova).lower()
    if 'master' in p: return 'Masters'
    if 'sub-18' in p or 'sub 18' in p or 'sub18' in p or 'u18' in p: return 'Sub18'
    if 'sub-16' in p or 'sub 16' in p or 'sub16' in p or 'u16' in p: return 'Sub16'
    if 'sub-14' in p or 'sub 14' in p or 'sub14' in p or 'u14' in p: return 'Sub14'
    
    if tipo_jogo == 'Nacional':
        if '2ª div' in p or '2a div' in p or 'ii div' in p or '2.ª div' in p: return 'CN2'
        if '1ª div' in p or '1a div' in p or 'i div' in p or 'senior' in p or 'sénior' in p: return 'CN1'
    else:
        if 'sub-21' in p or 'sub 21' in p or 'sub21' in p or 'u21' in p: return 'Sub21'
        if 'senior' in p or 'sénior' in p: return 'Seniores'
    return 'Outros'

def obter_preco_jogo(tipo_jogo, fase, genero, escalao):
    if MINHA_FUNCAO in TABELA_PREMIOS and tipo_jogo in TABELA_PREMIOS[MINHA_FUNCAO]:
        tabela_tipo = TABELA_PREMIOS[MINHA_FUNCAO][tipo_jogo]
        tabela_genero = tabela_tipo.get(genero, tabela_tipo.get("Masculino", {}))
        
        if tipo_jogo == "Associação":
            return tabela_genero.get(escalao, tabela_genero.get("Outros", 0.0))
        else:
            tabela_fase = tabela_genero.get(fase, tabela_genero.get("1ª Fase", {}))
            return tabela_fase.get(escalao, tabela_fase.get("Outros", 0.0))
    return 0.0

def obter_cidade(pavilhao):
    pav_limpo = str(pavilhao).lower()
    for palavra_chave, cidade in MAPA_PAVILHOES.items():
        if palavra_chave in pav_limpo: return cidade.strip()
    return pavilhao.strip()

def obter_dados_deslocacao(origem, pavilhao):
    origem_limpa = origem.strip()
    cidade_destino = obter_cidade(pavilhao)
    
    if DF_KMS is None or DF_EUROS is None:
        return 0, 0.0, cidade_destino
        
    try:
        kms = float(DF_KMS.at[origem_limpa, cidade_destino])
        valor_cru = str(DF_EUROS.at[origem_limpa, cidade_destino]).replace('€', '').replace(',', '.').strip()
        return kms, float(valor_cru), cidade_destino
    except Exception:
        return 0, 0.0, cidade_destino

# ==========================================
# FORMATADOR DO MAPA GERAL (AZUL)
# ==========================================
def formatar_folha_excel(ws, df_origem, titulo_principal, is_nacional=False):
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_cabecalho = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    fill_azul_claro = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws['A1'] = titulo_principal
    ws['A1'].font = title_font
    ws.merge_cells('A1:J1')
    
    ws['A2'] = "MAPA DE PRESTAÇÃO DE CONTAS - PRESTAÇÃO DE SERVIÇOS E/OU DESLOCAÇÕES"
    ws['A2'].font = bold_font
    ws.merge_cells('A2:J2')

    ws['A4'] = f"Nome: {NOME_COMPLETO}"
    ws['A5'] = f"Localidade: {LOCAL_PARTIDA}"
    ws['A6'] = f"IBAN: {MEU_IBAN}"
    
    ws['F4'] = f"Categoria: {MINHA_FUNCAO}"
    ws['J4'] = f"Mês: {MES_PROCESSAMENTO}/{ANO_PROCESSAMENTO}"
    
    for row in range(4, 7):
        for col in ['A', 'F', 'J']:
            if ws[f'{col}{row}'].value:
                ws[f'{col}{row}'].font = bold_font

    larguras = {'A': 12, 'B': 35, 'C': 12, 'D': 15, 'E': 25, 'F': 35, 'G': 8, 'H': 15, 'I': 15, 'J': 15}
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    max_row = ws.max_row
    
    if max_row >= 9:
        for row in ws.iter_rows(min_row=9, max_row=max_row, min_col=1, max_col=10):
            for cell in row:
                cell.border = thin_border
                
                if cell.row == 9:
                    cell.font = bold_font
                    cell.fill = fill_cabecalho
                    cell.alignment = Alignment(horizontal='center')
                elif cell.row > 9:
                    if cell.row % 2 == 0:
                        cell.fill = fill_azul_claro
                    if cell.column >= 7:
                        cell.alignment = Alignment(horizontal='center')

    total_deslocacoes = 0.0 if is_nacional else df_origem['Deslocacao Final'].sum() if not df_origem.empty else 0.0
    total_premios = df_origem['Premio Jogo'].sum() if not df_origem.empty else 0.0
    total_receber = total_premios + total_deslocacoes
    
    linha_total = max_row + 2
    ws[f'I{linha_total}'] = "TOTAL DESLOCAÇÕES:"
    ws[f'I{linha_total}'].font = bold_font
    ws[f'I{linha_total}'].alignment = Alignment(horizontal='right')
    ws[f'J{linha_total}'] = f"{total_deslocacoes:.2f} €"
    ws[f'J{linha_total}'].font = bold_font
    ws[f'J{linha_total}'].border = thin_border
    ws[f'J{linha_total}'].alignment = Alignment(horizontal='center')
    
    linha_total += 1
    ws[f'I{linha_total}'] = "TOTAL PRÉMIOS:"
    ws[f'I{linha_total}'].font = bold_font
    ws[f'I{linha_total}'].alignment = Alignment(horizontal='right')
    ws[f'J{linha_total}'] = f"{total_premios:.2f} €"
    ws[f'J{linha_total}'].font = bold_font
    ws[f'J{linha_total}'].border = thin_border
    ws[f'J{linha_total}'].alignment = Alignment(horizontal='center')

    linha_total += 1
    ws[f'I{linha_total}'] = "A RECEBER:"
    ws[f'I{linha_total}'].font = Font(bold=True, size=12)
    ws[f'I{linha_total}'].alignment = Alignment(horizontal='right')
    ws[f'J{linha_total}'] = f"{total_receber:.2f} €"
    ws[f'J{linha_total}'].font = Font(bold=True, size=12)
    ws[f'J{linha_total}'].border = thin_border
    ws[f'J{linha_total}'].alignment = Alignment(horizontal='center')
    ws[f'J{linha_total}'].fill = fill_cabecalho


# ==========================================
# O MOTOR DO SCRIPT
# ==========================================
def processar_tudo_automaticamente():
    print(f"A ligar ao site da FPB para o mes {MES_PROCESSAMENTO}/{ANO_PROCESSAMENTO}...")
    try:
        resposta = requests.get(URL_FPB, headers={'User-Agent': 'Mozilla/5.0'}, timeout=10)
        if resposta.status_code != 200: return print("Erro de acesso ao site.")
    except Exception as e:
        return print(f"Erro de conexão: {e}")

    try:
        tabelas = pd.read_html(StringIO(resposta.text))
        todos_os_jogos = pd.concat([tabelas[2], tabelas[3]], ignore_index=True)
    except Exception:
        return print("Erro: Nao encontrei as tabelas de jogos no site da FPB.")

    if isinstance(todos_os_jogos.columns, pd.MultiIndex):
        todos_os_jogos.columns = [c[-1] for c in todos_os_jogos.columns]
    todos_os_jogos.columns = [str(c).strip().replace('\xa0', ' ') for c in todos_os_jogos.columns]

    col_prova = next((c for c in todos_os_jogos.columns if 'prova' in c.lower() or 'competi' in c.lower()), None)
    col_data = next((c for c in todos_os_jogos.columns if 'data' in c.lower()), None)
    col_pavilhao = next((c for c in todos_os_jogos.columns if 'pavilh' in c.lower() or 'local' in c.lower() or 'recinto' in c.lower()), None)
    col_equipas = next((c for c in todos_os_jogos.columns if 'equipa' in c.lower() or 'jogo' in c.lower()), None)

    if not all([col_prova, col_data, col_pavilhao]):
        return print("Erro: Nao consegui identificar as colunas base.")

    todos_os_jogos['Data_Python'] = pd.to_datetime(todos_os_jogos[col_data], dayfirst=True, errors='coerce')
    jogos_do_mes = todos_os_jogos[
        (todos_os_jogos['Data_Python'].dt.month == MES_PROCESSAMENTO) &
        (todos_os_jogos['Data_Python'].dt.year == ANO_PROCESSAMENTO)
    ].copy()

    if jogos_do_mes.empty:
        return print(f"Nenhum jogo encontrado para {MES_PROCESSAMENTO}/{ANO_PROCESSAMENTO}.")

    # ==========================================
    # CALCULOS INTERNOS E AGRUPAMENTO
    # ==========================================
    jogos_do_mes = jogos_do_mes.sort_values(by='Data_Python')
    jogos_do_mes['Data_Apenas'] = jogos_do_mes['Data_Python'].dt.date
    
    jogos_do_mes['Tipo'] = jogos_do_mes[col_prova].apply(definir_tipo_competicao)
    jogos_do_mes['Genero'] = jogos_do_mes[col_prova].apply(extrair_genero)
    jogos_do_mes['Fase'] = jogos_do_mes[col_prova].apply(extrair_fase)
    jogos_do_mes['Escalao'] = jogos_do_mes.apply(lambda l: extrair_escalao(l[col_prova], l['Tipo']), axis=1)
    jogos_do_mes['Premio Jogo'] = jogos_do_mes.apply(lambda l: obter_preco_jogo(l['Tipo'], l['Fase'], l['Genero'], l['Escalao']), axis=1)

    novas_colunas = jogos_do_mes.apply(lambda l: obter_dados_deslocacao(LOCAL_PARTIDA, l[col_pavilhao]), axis=1, result_type='expand')
    jogos_do_mes['KMs'] = novas_colunas[0]
    jogos_do_mes['Valor Base'] = novas_colunas[1]
    jogos_do_mes['Cidade Destino'] = novas_colunas[2]

    jogos_do_mes['Paga Deslocacao'] = False
    jogos_do_mes['Deslocacao Final'] = 0.0
    jogos_do_mes['Pagador Deslocacao'] = ''

    grupos_dias = jogos_do_mes.groupby(['Data_Apenas', 'Cidade Destino'])
    for (data, cidade), grupo in grupos_dias:
        tem_nacional = 'Nacional' in grupo['Tipo'].values
        
        if tem_nacional:
            idx_nacional = grupo[grupo['Tipo'] == 'Nacional'].index[0]
            for idx in grupo.index:
                if idx == idx_nacional:
                    jogos_do_mes.at[idx, 'Paga Deslocacao'] = True
                    jogos_do_mes.at[idx, 'Deslocacao Final'] = jogos_do_mes.at[idx, 'Valor Base']
                    jogos_do_mes.at[idx, 'Pagador Deslocacao'] = 'PAGADOR'
                else:
                    jogos_do_mes.at[idx, 'Paga Deslocacao'] = False
                    jogos_do_mes.at[idx, 'Deslocacao Final'] = 0.0
                    jogos_do_mes.at[idx, 'Pagador Deslocacao'] = 'PAGO FPB'
        else:
            idx_primeiro = grupo.index[0]
            for idx in grupo.index:
                if idx == idx_primeiro:
                    jogos_do_mes.at[idx, 'Paga Deslocacao'] = True
                    jogos_do_mes.at[idx, 'Deslocacao Final'] = jogos_do_mes.at[idx, 'Valor Base']
                    jogos_do_mes.at[idx, 'Pagador Deslocacao'] = 'PAGADOR'
                else:
                    jogos_do_mes.at[idx, 'Paga Deslocacao'] = False
                    jogos_do_mes.at[idx, 'Deslocacao Final'] = 0.0
                    jogos_do_mes.at[idx, 'Pagador Deslocacao'] = 'PAGO ABA'

    # ==========================================
    # GERAR O MAPA GERAL (AZUL)
    # ==========================================
    jogos_nacional = jogos_do_mes[jogos_do_mes['Tipo'] == 'Nacional'].copy()
    jogos_associacao = jogos_do_mes[(jogos_do_mes['Tipo'] == 'Associação') & (jogos_do_mes['Escalao'] != 'Masters')].copy()
    jogos_masters = jogos_do_mes[(jogos_do_mes['Tipo'] == 'Associação') & (jogos_do_mes['Escalao'] == 'Masters')].copy()

    def formata_deslocacao(row, is_nacional):
        if row['Paga Deslocacao'] and row['Deslocacao Final'] > 0:
            return f"{row['Deslocacao Final']:.2f} €"
        elif not is_nacional and row['Pagador Deslocacao'] in ['PAGO FPB', 'PAGO ABA']:
            return row['Pagador Deslocacao']
        return ""

    def formatar_para_excel(df, is_nacional=False):
        return pd.DataFrame({
            "Data": df['Data_Python'].dt.strftime('%d/%m/%Y'),
            "Prova": df[col_prova],
            "Escalão": df['Escalao'],
            "Função": MINHA_FUNCAO,
            "Recinto": df[col_pavilhao],
            "Jogo": df[col_equipas] if col_equipas else "",
            "Kms": df['KMs'].apply(lambda x: x if x > 0 else ""),
            "Deslocação": df.apply(lambda row: formata_deslocacao(row, is_nacional), axis=1),
            "Prémio (€)": df['Premio Jogo'].apply(lambda x: f"{x:.2f} €"),
            "Total (€)": df.apply(lambda row: f"{row['Premio Jogo']:.2f} €" if is_nacional else f"{(row['Premio Jogo'] + row['Deslocacao Final']):.2f} €", axis=1)
        })

    caminho_base = f"{PASTA_DESTINO}\\Mapa_Contas_Geral_{MES_PROCESSAMENTO}_{ANO_PROCESSAMENTO}"
    caminho_final = f"{caminho_base}.xlsx"
    
    contador = 1
    while True:
        try:
            if os.path.exists(caminho_final):
                with open(caminho_final, 'a'): pass
            break
        except PermissionError:
            caminho_final = f"{caminho_base}_v{contador}.xlsx"
            contador += 1

    with pd.ExcelWriter(caminho_final, engine='openpyxl') as writer:
        escreveu_algo = False
        if not jogos_nacional.empty:
            formatar_para_excel(jogos_nacional, True).to_excel(writer, sheet_name="Nacional", index=False, startrow=8)
            escreveu_algo = True
        if not jogos_associacao.empty:
            formatar_para_excel(jogos_associacao, False).to_excel(writer, sheet_name="Associação", index=False, startrow=8)
            escreveu_algo = True
        if not jogos_masters.empty:
            formatar_para_excel(jogos_masters, False).to_excel(writer, sheet_name="Masters", index=False, startrow=8)
            escreveu_algo = True
        if not escreveu_algo:
            pd.DataFrame().to_excel(writer, sheet_name="Vazio")

    wb = openpyxl.load_workbook(caminho_final)
    if "Nacional" in wb.sheetnames:
        formatar_folha_excel(wb["Nacional"], jogos_nacional, "FEDERAÇÃO PORTUGUESA DE BASQUETEBOL", is_nacional=True)
    if "Associação" in wb.sheetnames:
        formatar_folha_excel(wb["Associação"], jogos_associacao, "ASSOCIAÇÃO DE BASQUETEBOL DE AVEIRO", is_nacional=False)
    if "Masters" in wb.sheetnames:
        formatar_folha_excel(wb["Masters"], jogos_masters, "ASSOCIAÇÃO DE AVEIRO - JOGOS MASTERS", is_nacional=False)

    try:
        folhas_existentes = wb.sheetnames
        if "Vazio" in folhas_existentes:
            if len(list(wb)) > 1:
                wb.remove(wb["Vazio"])
    except Exception:
        pass

    wb.save(caminho_final)

    # ==========================================
    # SEGUNDA FASE: PREENCHER OS TEMPLATES ORIGINAIS
    # ==========================================
    print("\n--- A PREENCHER TEMPLATES OFICIAIS DE DESPESAS ---")
    
    viagens_nacional = jogos_do_mes[(jogos_do_mes['Tipo'] == 'Nacional') & (jogos_do_mes['Paga Deslocacao'] == True)].copy()
    
    if not viagens_nacional.empty and os.path.exists(CAMINHO_TEMPLATE_NACIONAL):
        try:
            wb_nac = openpyxl.load_workbook(CAMINHO_TEMPLATE_NACIONAL)
            ws_nac = wb_nac.active
            
            for r in range(1, 10):
                for c in range(1, 10):
                    val = str(ws_nac.cell(row=r, column=c).value).strip().lower()
                    if 'mês:' in val or 'mes:' in val:
                        ws_nac.cell(row=r+1, column=c).value = MES_PROCESSAMENTO
                    if 'ano:' in val:
                        ws_nac.cell(row=r+1, column=c).value = ANO_PROCESSAMENTO
                        
            for _, row in viagens_nacional.iterrows():
                dia = int(row['Data_Python'].day)
                linha = 8 + dia 
                
                ws_nac[f'B{linha}'] = LOCAL_PARTIDA
                ws_nac[f'C{linha}'] = row['Cidade Destino']
                ws_nac[f'D{linha}'] = row[col_equipas] if col_equipas else row[col_prova]
                ws_nac[f'F{linha}'] = row['KMs']
                ws_nac[f'G{linha}'] = row['Valor Base']
                
            caminho_save_nac = os.path.join(PASTA_DESTINO, f"Despesas_Nacional_FPB_{MES_PROCESSAMENTO}_{ANO_PROCESSAMENTO}.xlsx")
            wb_nac.save(caminho_save_nac)
            print(f"-> Template Nacional FPB preenchido com sucesso: {caminho_save_nac}")
        except Exception as e:
            print(f"-> Erro ao preencher Template Nacional: {e}")
    else:
        if viagens_nacional.empty:
            print("-> Sem viagens Nacionais para faturar à FPB este mês.")
        else:
            print(f"-> Template Nacional ignorado: Não encontrei '{CAMINHO_TEMPLATE_NACIONAL}' na pasta.")

    viagens_aba = jogos_do_mes[(jogos_do_mes['Tipo'] == 'Associação')].copy()
    
    if not viagens_aba.empty and os.path.exists(CAMINHO_TEMPLATE_ABA):
        try:
            wb_aba = openpyxl.load_workbook(CAMINHO_TEMPLATE_ABA)
            ws_aba = wb_aba.active
            
            ws_aba['L65'] = MES_PROCESSAMENTO
            ws_aba['L67'] = ANO_PROCESSAMENTO

            linha_atual = 12
            for _, row in viagens_aba.iterrows():
                ws_aba[f'B{linha_atual}'] = row[col_pavilhao]
                ws_aba[f'C{linha_atual}'] = "" # Nº Jogo fica em branco
                ws_aba[f'D{linha_atual}'] = LOCAL_PARTIDA
                ws_aba[f'E{linha_atual}'] = row['Cidade Destino']
                
                if row['Paga Deslocacao']:
                    ws_aba[f'F{linha_atual}'] = row['KMs']
                    ws_aba[f'G{linha_atual}'] = row['Valor Base']
                else:
                    ws_aba[f'F{linha_atual}'] = 0
                    ws_aba[f'G{linha_atual}'] = 0
                    
                ws_aba[f'H{linha_atual}'] = row['Data_Python'].day
                ws_aba[f'J{linha_atual}'] = row['Data_Python'].day
                linha_atual += 1
                
            caminho_save_aba = os.path.join(PASTA_DESTINO, f"Despesas_Distrital_ABA_{MES_PROCESSAMENTO}_{ANO_PROCESSAMENTO}.xlsx")
            wb_aba.save(caminho_save_aba)
            print(f"-> Template Distrital ABA preenchido com sucesso: {caminho_save_aba}")
        except Exception as e:
            print(f"-> Erro ao preencher Template Distrital ABA: {e}")
    else:
        if viagens_aba.empty:
            print("-> Sem viagens Distritais para faturar à ABA este mês.")
        else:
            print(f"-> Template Distrital ignorado: Não encontrei '{CAMINHO_TEMPLATE_ABA}' na pasta.")

    print("\n" + "="*50)
    print("PROCESSO TOTAL CONCLUÍDO COM SUCESSO!")
    print("="*50)

processar_tudo_automaticamente()