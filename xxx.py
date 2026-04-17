from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import os
import json
import html 
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import Counter
from datetime import datetime

# ==========================================
# CONFIGURAÇÕES DE ACESSO E DADOS
# ==========================================
NOME_UTILIZADOR = "gabrielcorreiadelfim@gmail.com"  
PASSWORD = "#Areiased5"      
URL_LOGIN = "https://juizesbasquetebol.pt/login"  
URL_JOGOS = "https://juizesbasquetebol.pt/dashboard/os-meus-jogos/jogos-realizados" 

MEU_NOME_SISTEMA = "Delfim Gabriel Almeida Correia"
MEU_NOME_CURTO = "Delfim" 

# Dados para o Mapa de Contas
MEU_IBAN = "PT50 0000 0000 0000 0000 0" 
LOCAL_PARTIDA = "Vale de Cambra"

PASTA_DESTINO = r"C:\Users\Gabriel\Documents\vscode\scrpit"
FICHEIRO_FINAL = os.path.join(PASTA_DESTINO, "Dashboard_Arbitragem.xlsx")
CAMINHO_TABELA_KMS = os.path.join(PASTA_DESTINO, "Tabela_KMs.xlsx")
CAMINHO_TABELA_EUROS = os.path.join(PASTA_DESTINO, "Tabela_Euros.xlsx")

# ==========================================
# TABELAS DE PREÇOS E PAVILHÕES
# ==========================================
# SE ALGUM PRÉMIO ESTIVER ERRADO, MUDA OS VALORES AQUI:
TABELA_PREMIOS = {
    "Oficial de Mesa": {
        "Nacional": {
            "Masculino": {"1ª Fase": {"CN1": 25.0, "CN2": 24.0, "Sub18": 19.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0},
                          "2ª Fase": {"CN1": 26.0, "CN2": 25.0, "Sub18": 19.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0}},
            "Feminino": {"1ª Fase": {"CN1": 19.0, "CN2": 17.0, "Sub18": 17.0, "Sub16": 15.0, "Sub14": 15.0, "Outros": 0},
                         "2ª Fase": {"CN1": 20.0, "CN2": 18.0, "Sub18": 17.0, "Sub16": 15.0, "Sub14": 15.0, "Outros": 0}}
        },
        "Associação": {
            "Masculino": {"Seniores": 19.0, "Masters": 14.0, "Sub21": 12.0, "Sub18": 10.0, "Sub16": 9.0, "Sub14": 8.0, "Outros": 0},
            "Feminino": {"Seniores": 19.0, "Masters": 14.0, "Sub21": 12.0, "Sub18": 10.0, "Sub16": 9.0, "Sub14": 8.0, "Outros": 0}
        }
    },
    "Arbitro": {
        "Nacional": {
            "Masculino": {"1ª Fase": {"CN1": 46.0, "CN2": 35.0, "Sub18": 28.0, "Sub16": 25.0, "Sub14": 25.0, "Outros": 0},
                          "2ª Fase": {"CN1": 47.0, "CN2": 38.0, "Sub18": 26.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0}},
            "Feminino": {"1ª Fase": {"CN1": 34.0, "CN2": 24.0, "Sub18": 23.0, "Sub16": 20.0, "Sub14": 18.0, "Outros": 0},
                         "2ª Fase": {"CN1": 36.0, "CN2": 25.0, "Sub18": 23.0, "Sub16": 20.0, "Sub14": 18.0, "Outros": 0}}
        },
        "Associação": {
            "Masculino": {"Seniores": 27.0, "Masters": 27.0, "Sub21": 14.0, "Sub18": 13.0, "Sub16": 11.0, "Sub14": 10.0, "Outros": 0},
            "Feminino": {"Seniores": 27.0, "Masters": 27.0, "Sub21": 14.0, "Sub18": 13.0, "Sub16": 11.0, "Sub14": 10.0, "Outros": 0}
        }
    }
}

# SE UM PAVILHÃO ESTIVER A DAR 0 KMs, ADICIONA UMA PALAVRA-CHAVE DELE AQUI:
MAPA_PAVILHOES = {
    "arena de ovar": "Ovar", "ovar": "Ovar",
    "municipal de anadia": "Anadia", "acr cerca": "Anadia", "anadia": "Anadia",
    "ventosa do bairro": "Mealhada", "mealhada": "Mealhada",
    "vale cambra": "Vale de Cambra", "vale de cambra": "Vale de Cambra",
    "paulo pinto": "S.João da Madeira", "são joão da madeira": "S.João da Madeira",
    "municipal de vagos": "Vagos", "vagos": "Vagos",
    "oliveira do bairro": "Oliveira do Bairro",
    "galitos": "Aveiro", "joão afonso": "Aveiro", "aveiro": "Aveiro",
    "secundária de arouca": "Arouca", "arouca": "Arouca",
    "albergaria": "Albergaria a Velha",
    "esgueira": "Esgueira", "costeira": "Oliveira de Azeméis", "salvador machado": "Oliveira de Azeméis",
    "oliveira de azeméis": "Oliveira de Azeméis", "águeda": "Águeda", "gica": "Águeda",
    "gafanha": "Gafanha da Nazaré", "antónio júlio silva": "Paços de Brandão", "paços de brandão": "Paços de Brandão",
    "adriano nordeste": "Ílhavo", "illiabum": "Ílhavo", "ílhavo": "Ílhavo",
    "estarreja": "Estarreja", "sangalhos": "Sangalhos", "luso": "Luso", "calvão": "Calvão"
}

# ==========================================
# FUNÇÕES DE LÓGICA E CÁLCULO
# ==========================================
def carregar_matrizes_kms():
    df_kms, df_euros = None, None
    try:
        if os.path.exists(CAMINHO_TABELA_KMS) and os.path.exists(CAMINHO_TABELA_EUROS):
            df_kms = pd.read_excel(CAMINHO_TABELA_KMS, index_col=0)
            df_euros = pd.read_excel(CAMINHO_TABELA_EUROS, index_col=0)
            df_kms.index, df_kms.columns = df_kms.index.astype(str).str.strip(), df_kms.columns.astype(str).str.strip()
            df_euros.index, df_euros.columns = df_euros.index.astype(str).str.strip(), df_euros.columns.astype(str).str.strip()
    except Exception as e:
        pass
    return df_kms, df_euros

DF_KMS, DF_EUROS = carregar_matrizes_kms()

def obter_cidade(pavilhao):
    pav_limpo = str(pavilhao).lower()
    for chave, cidade in MAPA_PAVILHOES.items():
        if chave in pav_limpo: return cidade
    return pavilhao.strip()

def obter_dados_deslocacao(origem, pavilhao):
    origem_limpa = origem.strip()
    cidade = obter_cidade(pavilhao)
    if DF_KMS is None or DF_EUROS is None: return 0.0, 0.0, cidade
    try:
        kms = float(DF_KMS.at[origem_limpa, cidade])
        euros = float(str(DF_EUROS.at[origem_limpa, cidade]).replace('€','').replace(',','.').strip())
        return kms, euros, cidade
    except:
        return 0.0, 0.0, cidade

def classificar_jogo(row):
    comp = str(row.get('Competição', '')).lower()
    tipo = 'Associação'
    if 'master' in comp: tipo = 'Masters'
    elif 'nacional' in comp or 'festa do basquetebol' in comp or 'lpb' in comp or 'proliga' in comp: tipo = 'Nacional'
    
    gen = 'Feminino' if 'fem' in comp else 'Masculino'
    fase = '2ª Fase' if '2ª fase' in str(row.get('Fase', '')).lower() or 'final' in str(row.get('Fase', '')).lower() else '1ª Fase'
    
    esc = 'Outros'
    if 'sub 18' in comp or 'sub18' in comp or 'u18' in comp: esc = 'Sub18'
    elif 'sub 16' in comp or 'sub16' in comp or 'u16' in comp: esc = 'Sub16'
    elif 'sub 14' in comp or 'sub14' in comp or 'u14' in comp: esc = 'Sub14'
    elif 'sub 21' in comp or 'sub21' in comp or 'u21' in comp: esc = 'Sub21'
    elif 'senior' in comp or 'sénior' in comp or 'lpb' in comp or 'proliga' in comp or '1ª div' in comp or '2ª div' in comp: esc = 'Seniores'
    elif 'master' in comp: esc = 'Masters'
    
    if tipo == 'Nacional' and esc == 'Seniores':
        esc = 'CN2' if '2ª div' in comp else 'CN1'

    return pd.Series([tipo, gen, fase, esc])

def calcular_premios_e_viagens(df):
    if df.empty: return df
    
    df[['Tipo', 'Genero', 'Fase_Calculo', 'Escalao_Calculo']] = df.apply(classificar_jogo, axis=1)
    
    premios = []
    for _, row in df.iterrows():
        categoria = "Oficial de Mesa" if "Árbitro" not in str(row['Minha Função']) else "Arbitro"
        tipo, gen, fase, esc = row['Tipo'], row['Genero'], row['Fase_Calculo'], row['Escalao_Calculo']
        
        tipo_busca = 'Associação' if tipo == 'Masters' else tipo
        
        try:
            t_tipo = TABELA_PREMIOS.get(categoria, {}).get(tipo_busca, {})
            t_gen = t_tipo.get(gen, t_tipo.get("Masculino", {}))
            if tipo_busca == "Associação": val = t_gen.get(esc, 0.0)
            else: val = t_gen.get(fase, {}).get(esc, 0.0)
        except: val = 0.0
        premios.append(val)
    df['Prémio (€)'] = premios

    # Mês do Jogo
    df['Data_Date'] = pd.to_datetime(df['Data'], format="%d-%m-%Y", errors='coerce')
    df['Mês_Ano'] = df['Data_Date'].dt.strftime('%m-%Y')

    deslocacoes, kms_list, cids = [], [], []
    for _, row in df.iterrows():
        kms, euros, cid = obter_dados_deslocacao(LOCAL_PARTIDA, row['Recinto'])
        kms_list.append(kms)
        deslocacoes.append(euros)
        cids.append(cid)
        
    df['KMs'] = kms_list
    df['Valor Viagem'] = deslocacoes
    df['Cidade Destino'] = cids
    df['Deslocação (€)'] = 0.0

    # Agrupa por Dia e Cidade Destino para não pagar a viagem 2x no mesmo dia
    for (data, cid), grupo in df.groupby(['Data_Date', 'Cidade Destino']):
        tem_nac = 'Nacional' in grupo['Tipo'].values
        idx_paga = grupo[grupo['Tipo'] == 'Nacional'].index[0] if tem_nac else grupo.index[0]
        
        for idx in grupo.index:
            if idx == idx_paga: df.at[idx, 'Deslocação (€)'] = df.at[idx, 'Valor Viagem']
            else: df.at[idx, 'Deslocação (€)'] = 0.0

    df['Total (€)'] = df['Prémio (€)'] + df['Deslocação (€)']
    return df

def gerar_resumo_financeiro(df):
    if df.empty: return pd.DataFrame()
    
    resumo = []
    meses_unicos = df['Mês_Ano'].dropna().unique()
    meses_unicos = sorted(meses_unicos, key=lambda x: datetime.strptime(x, '%m-%Y'))

    for mes in meses_unicos:
        df_mes = df[df['Mês_Ano'] == mes]
        
        g_desl = df_mes['Deslocação (€)'].sum()
        g_prem = df_mes['Prémio (€)'].sum()
        g_tot = df_mes['Total (€)'].sum()
        
        df_aba = df_mes[df_mes['Tipo'] == 'Associação']
        a_desl = df_aba['Deslocação (€)'].sum()
        a_prem = df_aba['Prémio (€)'].sum()
        a_tot = df_aba['Total (€)'].sum()
        
        df_nac = df_mes[df_mes['Tipo'] == 'Nacional']
        n_desl = df_nac['Deslocação (€)'].sum()
        n_prem = df_nac['Prémio (€)'].sum()
        n_tot = df_nac['Total (€)'].sum()
        
        df_mas = df_mes[df_mes['Tipo'] == 'Masters']
        m_desl = df_mas['Deslocação (€)'].sum()
        m_prem = df_mas['Prémio (€)'].sum()
        m_tot = df_mas['Total (€)'].sum()
        
        resumo.append({
            "Mês": mes,
            "Geral - Deslocação": g_desl, "Geral - Prémio": g_prem, "Geral - Total": g_tot,
            "ABA - Deslocação": a_desl, "ABA - Prémio": a_prem, "ABA - Total": a_tot,
            "Nacional - Deslocação": n_desl, "Nacional - Prémio": n_prem, "Nacional - Total": n_tot,
            "Masters - Deslocação": m_desl, "Masters - Prémio": m_prem, "Masters - Total": m_tot
        })
        
    return pd.DataFrame(resumo)

# ==========================================
# FUNÇÕES DE FORMATAÇÃO DE EXCEL (OPENPYXL)
# ==========================================
def formatar_geral(ws):
    font_titulo = Font(name='Segoe UI', size=16, bold=True, color="FFFFFF")
    font_header = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    border_all = Border(top=Side(style='thin', color="BFBFBF"), left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells('A1:I2')
    ws['A1'] = "BASE DE DADOS GLOBAL - JOGOS"
    ws['A1'].font = font_titulo
    ws['A1'].fill = PatternFill(start_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = align_center

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = font_header
        cell.fill = PatternFill(start_color="2F75B5", fill_type="solid")
        cell.alignment = align_center
        cell.border = border_all
    
    ws.auto_filter.ref = f"A4:{openpyxl.utils.get_column_letter(ws.max_column)}4"
    ws.freeze_panes = "A5"

    for r in range(5, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all
            cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid") if r % 2 == 0 else PatternFill(start_color="FFFFFF", fill_type="solid")
            if c in [1, 2, 3, 9]: cell.alignment = align_center
            else: cell.alignment = align_left

    larguras = {1:13, 2:8, 3:10, 4:12, 5:35, 6:40, 7:35, 8:18, 9:20, 10:25, 11:25, 12:25, 13:25, 14:25, 15:25, 16:25}
    for c, w in larguras.items():
        if c <= ws.max_column: ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

def formatar_mapa_contas(ws, titulo_entidade):
    font_titulo = Font(name='Segoe UI', size=14, bold=True)
    font_bold = Font(name='Segoe UI', bold=True)
    border_all = Border(top=Side(style='thin', color="000000"), left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))
    fill_cabecalho = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    fill_zebra = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws.merge_cells('A1:K1')
    ws['A1'] = titulo_entidade
    ws['A1'].font = font_titulo
    
    ws.merge_cells('A2:K2')
    ws['A2'] = "MAPA DE PRESTAÇÃO DE CONTAS - PRESTAÇÃO DE SERVIÇOS E/OU DESLOCAÇÕES"
    ws['A2'].font = font_bold
    
    ws['A4'] = f"Nome: {MEU_NOME_SISTEMA}"
    ws['A5'] = f"Localidade: {LOCAL_PARTIDA}"
    ws['A6'] = f"IBAN: {MEU_IBAN}"
    ws['F4'] = f"Categoria: Oficial de Mesa / Árbitro"
    
    for row in range(4, 7):
        for col in ['A', 'F', 'K']:
            if ws[f'{col}{row}'].value: ws[f'{col}{row}'].font = font_bold

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=8, column=col)
        cell.font = font_bold
        cell.fill = fill_cabecalho
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all

    ws.auto_filter.ref = f"A8:{openpyxl.utils.get_column_letter(ws.max_column)}8"
    ws.freeze_panes = "A9"

    for r in range(9, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all
            if r % 2 == 0: cell.fill = fill_zebra
            if c >= 8: cell.alignment = Alignment(horizontal="center")
            
            if c in [9, 10, 11] and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 €'

    larguras = {1:12, 2:12, 3:35, 4:12, 5:20, 6:35, 7:40, 8:8, 9:15, 10:15, 11:15}
    for c, w in larguras.items():
        if c <= ws.max_column: ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
        
    max_r = ws.max_row
    linha_totais = max_r + 2 
    
    ws[f'J{linha_totais}'] = "TOTAL DESLOCAÇÕES:"
    ws[f'J{linha_totais}'].font = font_bold
    ws[f'J{linha_totais}'].alignment = Alignment(horizontal="right")
    ws[f'K{linha_totais}'] = f"=SUBTOTAL(109, I9:I{max_r})" 
    ws[f'K{linha_totais}'].font = font_bold
    ws[f'K{linha_totais}'].number_format = '#,##0.00 €'
    
    ws[f'J{linha_totais+1}'] = "TOTAL PRÉMIOS:"
    ws[f'J{linha_totais+1}'].font = font_bold
    ws[f'J{linha_totais+1}'].alignment = Alignment(horizontal="right")
    ws[f'K{linha_totais+1}'] = f"=SUBTOTAL(109, J9:J{max_r})"
    ws[f'K{linha_totais+1}'].font = font_bold
    ws[f'K{linha_totais+1}'].number_format = '#,##0.00 €'
    
    ws[f'J{linha_totais+2}'] = "A RECEBER:"
    ws[f'J{linha_totais+2}'].font = Font(name='Segoe UI', size=12, bold=True)
    ws[f'J{linha_totais+2}'].alignment = Alignment(horizontal="right")
    ws[f'K{linha_totais+2}'] = f"=SUBTOTAL(109, K9:K{max_r})"
    ws[f'K{linha_totais+2}'].font = Font(name='Segoe UI', size=12, bold=True)
    ws[f'K{linha_totais+2}'].number_format = '#,##0.00 €'
    ws[f'K{linha_totais+2}'].fill = fill_cabecalho

def formatar_estatisticas_gerais(ws, is_resumo=False):
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F2F2", fill_type="solid")
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center")
            else:
                if cell.row % 2 == 0: cell.fill = fill_zebra
                
            # Formatação de Euros APENAS no Resumo Financeiro
            if is_resumo:
                if "Total" in str(ws.cell(row=1, column=cell.column).value) or "Prémio" in str(ws.cell(row=1, column=cell.column).value) or "Deslocação" in str(ws.cell(row=1, column=cell.column).value):
                    if cell.row > 1 and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00 €'
            else:
                cell.alignment = Alignment(horizontal="left")

    larguras = {1:15, 2:20, 3:20, 4:20, 5:20, 6:20, 7:20, 8:20, 9:20, 10:20, 11:20, 12:20, 13:20}
    if not is_resumo:
        larguras = {1:30, 2:15, 4:30, 5:15, 7:40, 8:15} # Larguras para a aba de estatísticas
        
    for c, w in larguras.items():
        if c <= ws.max_column: ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

# ==========================================
# MOTOR DO EXTRATOR
# ==========================================
def extrair_gerar_dashboard():
    jogos_conhecidos = set()
    df_antigo = pd.DataFrame()
    
    if os.path.exists(FICHEIRO_FINAL):
        try:
            df_teste = pd.read_excel(FICHEIRO_FINAL, sheet_name="Geral", nrows=5)
            if "HISTÓRICO" in str(df_teste.columns[0]) or "BASE" in str(df_teste.columns[0]):
                df_antigo = pd.read_excel(FICHEIRO_FINAL, sheet_name="Geral", skiprows=3)
            else:
                df_antigo = pd.read_excel(FICHEIRO_FINAL, sheet_name="Geral")
            
            if 'Nº Jogo' in df_antigo.columns:
                jogos_conhecidos = set(df_antigo['Nº Jogo'].astype(str).tolist())
                print(f"-> Excel detetado com {len(jogos_conhecidos)} jogos. A iniciar sincronização rápida...")
        except Exception as e:
            print(f"-> Aviso ao ler histórico antigo: A criar ficheiro novo do zero.")

    print("A iniciar o robô do Chrome...")
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 15) 
    todos_os_jogos = [] 
    ja_existe = False
    
    try:
        driver.get(URL_LOGIN)
        wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(NOME_UTILIZADOR)
        driver.find_element(By.ID, "password").send_keys(PASSWORD)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(5) 
        
        driver.get(URL_JOGOS)
        time.sleep(4)
        
        pagina = 1
        while True:
            print(f"A extrair Página {pagina}...")
            wait.until(lambda d: d.execute_script("return typeof Fabrik !== 'undefined' && Fabrik.blocks['list_5_com_fabrik_5'] !== undefined;"))
            time.sleep(1)
            
            dados_memoria = driver.execute_script("return Fabrik.blocks['list_5_com_fabrik_5'].options.data;")
            if not dados_memoria or len(dados_memoria[0]) == 0: break
                
            for jogo in dados_memoria[0]:
                row = jogo['data']
                n_jogo = str(row.get('fab_jogos___num_jogo_raw', ''))
                
                if n_jogo in jogos_conhecidos:
                    print(f"\n>>> Jogo {n_jogo} já existe. Sincronização concluída!")
                    ja_existe = True
                    break
                
                comp_limpa = html.unescape(row.get('fab_jogos___competicao', ''))
                fase_limpa = html.unescape(row.get('fab_jogos___fase', ''))
                recinto_limpo = html.unescape(row.get('fab_jogos___recinto', ''))
                visitado = html.unescape(row.get('fab_jogos___visitado', '').strip())
                visitante = html.unescape(row.get('fab_jogos___visitante', '').strip())
                
                juizes = json.loads(row.get('fab_nomeacoes___juiz', '[]'))
                funcoes = json.loads(row.get('fab_nomeacoes___funcao_jogo', '[]'))
                
                minha_f = "N/A"
                equipa_dict = {
                    "Árbitro Principal": "", "Árbitro Auxiliar 1": "", "Árbitro Auxiliar 2": "",
                    "Marcador": "", "Marcador Auxiliar": "", "Cronometrista": "", "Operador 24\"": ""
                }
                
                for nome, funcao in zip(juizes, funcoes):
                    nome_limpo = html.unescape(nome)
                    if MEU_NOME_CURTO.lower() in nome_limpo.lower(): minha_f = funcao
                    if funcao in equipa_dict: equipa_dict[funcao] = nome_limpo
                    else: equipa_dict["Árbitro Auxiliar 2"] += f"{nome_limpo} "
                
                todos_os_jogos.append({
                    "Data": row.get('fab_jogos___data', ''),
                    "Hora": row.get('fab_jogos___hora', ''),
                    "Nº Jogo": n_jogo,
                    "Recinto": recinto_limpo,
                    "Equipas": f"{visitado} vs {visitante}",
                    "Competição": comp_limpa,
                    "Fase": fase_limpa,
                    "Minha Função": minha_f,
                    "Árbitro Principal": equipa_dict["Árbitro Principal"],
                    "Árbitro Auxiliar 1": equipa_dict["Árbitro Auxiliar 1"],
                    "Árbitro Auxiliar 2": equipa_dict["Árbitro Auxiliar 2"],
                    "Marcador": equipa_dict["Marcador"],
                    "Marcador Auxiliar": equipa_dict["Marcador Auxiliar"],
                    "Cronometrista": equipa_dict["Cronometrista"],
                    "Operador 24\"": equipa_dict["Operador 24\""]
                })

            if ja_existe: break
                
            try:
                btn = driver.find_element(By.XPATH, "//a[@rel='next' or contains(text(), '»')]")
                url_next = btn.get_attribute("href")
                if not url_next or url_next.endswith("#") or "disabled" in btn.find_element(By.XPATH, "..").get_attribute("class"): break
                driver.get(url_next)
                time.sleep(4)
                pagina += 1
            except: break

        # ==========================================
        # PROCESSAMENTO DE DADOS (PANDAS)
        # ==========================================
        if todos_os_jogos or not df_antigo.empty:
            print("\nA calcular dados financeiros e a gerar os separadores do Dashboard...")
            
            if todos_os_jogos:
                df_novos = pd.DataFrame(todos_os_jogos)
                df_completo = pd.concat([df_novos, df_antigo], ignore_index=True)
                df_completo = df_completo.drop_duplicates(subset=['Nº Jogo'], keep='first')
            else:
                df_completo = df_antigo.copy()

            df_completo = calcular_premios_e_viagens(df_completo)
            
            df_resumo_financeiro = gerar_resumo_financeiro(df_completo)
            
            colunas_mapa = ["Mês_Ano", "Data", "Competição", "Escalao_Calculo", "Minha Função", "Recinto", "Equipas", "KMs", "Deslocação (€)", "Prémio (€)", "Total (€)"]
            
            df_nac = df_completo[df_completo['Tipo'] == 'Nacional'][colunas_mapa].copy()
            df_assoc = df_completo[df_completo['Tipo'] == 'Associação'][colunas_mapa].copy()
            df_mast = df_completo[df_completo['Tipo'] == 'Masters'][colunas_mapa].copy()

            renomes = {"Mês_Ano": "Mês", "Competição": "Prova", "Escalao_Calculo": "Escalão", "Minha Função": "Função", "Equipas": "Jogo"}
            df_nac.rename(columns=renomes, inplace=True)
            df_assoc.rename(columns=renomes, inplace=True)
            df_mast.rename(columns=renomes, inplace=True)

            stats_funcao = df_completo['Minha Função'].value_counts().reset_index()
            stats_funcao.columns = ['Função', 'Total Jogos']
            stats_recinto = df_completo['Recinto'].value_counts().reset_index()
            stats_recinto.columns = ['Pavilhão', 'Total Jogos']
            stats_comp = df_completo['Competição'].value_counts().reset_index()
            stats_comp.columns = ['Competição', 'Total Jogos']

            cols_equipa = ["Árbitro Principal", "Árbitro Auxiliar 1", "Árbitro Auxiliar 2", "Marcador", "Marcador Auxiliar", "Cronometrista", "Operador 24\""]
            colegas = []
            for col in cols_equipa:
                for val in df_completo[col].dropna():
                    if val and val.strip() and MEU_NOME_CURTO.lower() not in val.lower(): colegas.append(val.strip())
            df_equipa = pd.DataFrame(Counter(colegas).most_common(), columns=['Colega de Equipa', 'Nº de Jogos Trabalhados'])

            with pd.ExcelWriter(FICHEIRO_FINAL, engine='openpyxl') as writer:
                if not df_resumo_financeiro.empty:
                    df_resumo_financeiro.to_excel(writer, sheet_name="Resumo Financeiro", index=False)
                
                df_base = df_completo.drop(columns=['Tipo', 'Genero', 'Fase_Calculo', 'Escalao_Calculo', 'Data_Date', 'Cidade Destino', 'Valor Viagem', 'Mês_Ano'], errors='ignore')
                df_base.to_excel(writer, sheet_name="Geral", index=False, startrow=3)
                
                df_nac.to_excel(writer, sheet_name="Nacional", index=False, startrow=7)
                df_assoc.to_excel(writer, sheet_name="Associação", index=False, startrow=7)
                if not df_mast.empty: df_mast.to_excel(writer, sheet_name="Masters", index=False, startrow=7)
                
                stats_funcao.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=0)
                stats_recinto.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=3)
                stats_comp.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=6)
                
                df_equipa.to_excel(writer, sheet_name="Minha Equipa", index=False)

            wb = openpyxl.load_workbook(FICHEIRO_FINAL)
            
            if "Resumo Financeiro" in wb.sheetnames: formatar_estatisticas_gerais(wb["Resumo Financeiro"], is_resumo=True)
            formatar_geral(wb["Geral"])
            formatar_mapa_contas(wb["Nacional"], "FEDERAÇÃO PORTUGUESA DE BASQUETEBOL")
            formatar_mapa_contas(wb["Associação"], "ASSOCIAÇÃO DE BASQUETEBOL DE AVEIRO")
            if "Masters" in wb.sheetnames: formatar_mapa_contas(wb["Masters"], "ASSOCIAÇÃO DE AVEIRO - JOGOS MASTERS")
            
            for ws_name in ["Estatísticas", "Minha Equipa"]:
                if ws_name in wb.sheetnames: formatar_estatisticas_gerais(wb[ws_name], is_resumo=False)

            wb.save(FICHEIRO_FINAL)
            print(f"\n✅ SUCESSO! Dashboard completo criado e atualizado.")
            print(f"Local: {FICHEIRO_FINAL}")
            
    except Exception as e: print(f"Erro: {e}")
    finally: driver.quit()

if __name__ == "__main__":
    extrair_gerar_dashboard()