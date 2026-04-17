import requests
import pandas as pd
from io import StringIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# ==========================================
# CONFIGURACOES PRINCIPAIS
# ==========================================
MES_PROCESSAMENTO = 11
ANO_PROCESSAMENTO = 2025
NOME_COMPLETO = "Gabriel [O Teu Apelido]"  # PREENCHE AQUI
MEU_IBAN = "PT50 0000 0000 0000 0000 0"    # PREENCHE AQUI
LOCAL_PARTIDA = "Vale de Cambra"
MINHA_FUNCAO = "Oficial de Mesa" 
FASE_ATUAL = "1FASE" # Escreve "1FASE" ou "2FASE"
PASTA_DESTINO = r"C:\Users\Gabriel\Documents\vscode\scrpit"
URL_FPB = "https://www.fpb.pt/juizes/18534/"

# Caminhos para as tuas tabelas de base de dados
CAMINHO_TABELA_KMS = os.path.join(PASTA_DESTINO, "Tabela_KMs.xlsx")
CAMINHO_TABELA_EUROS = os.path.join(PASTA_DESTINO, "Tabela_Euros.xlsx")

# ==========================================
# 1. TABELAS DE PREMIOS
# ==========================================
TABELA_PREMIOS = {
    "Oficial de Mesa": {
        "Nacional": {
            "Masculino": {
                "1ª Fase": { "CN1": 25.0, "CN2": 24.0, "Sub18": 19.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0 },
                "2ª Fase": { "CN1": 26.0, "CN2": 25.0, "Sub18": 19.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0 }
            },
            "Feminino": {
                "1ª Fase": { "CN1": 19.0, "CN2": 17.0, "Sub18": 17.0, "Sub16": 15.0, "Sub14": 15.0, "Outros": 0 },
                "2ª Fase": { "CN1": 20.0, "CN2": 18.0, "Sub18": 17.0, "Sub16": 15.0, "Sub14": 15.0, "Outros": 0 }
            }
        },
        "Associação": {
            "Masculino": { "Seniores": 19.0, "Masters": 19.0, "Sub21": 12.0, "Sub18": 10.0, "Sub16": 9.0, "Sub14": 8.0, "Outros": 0 },
            "Feminino": { "Seniores": 19.0, "Masters": 19.0, "Sub21": 12.0, "Sub18": 10.0, "Sub16": 9.0, "Sub14": 8.0, "Outros": 0 }
        }
    },
    "Arbitro": {
        "Nacional": {
            "Masculino": {
                "1ª Fase": { "CN1": 46.0, "CN2": 35.0, "Sub18": 28.0, "Sub16": 25.0, "Sub14": 25.0, "Outros": 0 },
                "2ª Fase": { "CN1": 47.0, "CN2": 38.0, "Sub18": 26.0, "Sub16": 16.0, "Sub14": 14.0, "Outros": 0 }
            },
            "Feminino": {
                "1ª Fase": { "CN1": 34.0, "CN2": 24.0, "Sub18": 23.0, "Sub16": 20.0, "Sub14": 18.0, "Outros": 0 },
                "2ª Fase": { "CN1": 36.0, "CN2": 25.0, "Sub18": 23.0, "Sub16": 20.0, "Sub14": 18.0, "Outros": 0 }
            }
        },
        "Associação": {
            "Masculino": { "Seniores": 27.0, "Masters": 27.0, "Sub21": 14.0, "Sub18": 13.0, "Sub16": 11.0, "Sub14": 10.0, "Outros": 0 },
            "Feminino": { "Seniores": 27.0, "Masters": 27.0, "Sub21": 14.0, "Sub18": 13.0, "Sub16": 11.0, "Sub14": 10.0, "Outros": 0 }
        }
    }
}

MAPA_PAVILHOES = {
    "arena de ovar": "Ovar",
    "ovar": "Ovar",
    "municipal de anadia": "Anadia",
    "acr cerca": "Anadia",
    "s. pedro": "Anadia",
    "anadia": "Anadia",
    "ventosa do bairro": "Mealhada",
    "mealhada": "Mealhada",
    "vale cambra": "Vale de Cambra",
    "vale de cambra": "Vale de Cambra",
    "paulo pinto": "S.João da Madeira",
    "são joão da madeira": "S.João da Madeira",
    "sao joao da madeira": "S.João da Madeira",
    "s. joão": "S.João da Madeira",
    "s. joao": "S.João da Madeira",
    "municipal de vagos": "Vagos",
    "vagos": "Vagos",
    "oliveira do bairro": "Oliveira do Bairro",
    "galitos": "Aveiro",
    "joão afonso": "Aveiro",
    "joao afonso": "Aveiro",
    "aveiro": "Aveiro",
    "secundária de arouca": "Arouca",
    "secundaria de arouca": "Arouca",
    "arouca": "Arouca",
    "albergaria": "Albergaria a Velha",
    "esgueira": "Esgueira",
    "clube do povo de esgueira": "Esgueira",
    "costeira": "Oliveira de Azeméis",
    "antónio costeira": "Oliveira de Azeméis",
    "antonio costeira": "Oliveira de Azeméis",
    "salvador machado": "Oliveira de Azeméis",
    "oliveira de azeméis": "Oliveira de Azeméis",
    "oliveira de azemeis": "Oliveira de Azeméis",
    "ferreira castro": "Oliveira de Azeméis",
    "águeda": "Águeda",
    "agueda": "Águeda",
    "gica": "Águeda",
    "gafanha": "Gafanha da Nazaré",
    "antónio júlio silva": "Paços de Brandão",
    "antonio julio silva": "Paços de Brandão",
    "paços de brandão": "Paços de Brandão",
    "pacos de brandao": "Paços de Brandão",
    "adriano nordeste": "Ílhavo",
    "illiabum": "Ílhavo",
    "ílhavo": "Ílhavo",
    "ilhavo": "Ílhavo",
    "estarreja": "Estarreja",
    "sangalhos": "Sangalhos",
    "complexo desportivo de sangalhos": "Sangalhos",
    "luso": "Luso",
    "calvão": "Calvão",
    "calvao": "Calvão"
}

# ==========================================
# CARREGAR BASES DE DADOS
# ==========================================
DF_KMS = None
DF_EUROS = None
caminho_csv_kms = CAMINHO_TABELA_KMS.replace(".xlsx", ".csv")
caminho_csv_euros = CAMINHO_TABELA_EUROS.replace(".xlsx", ".csv")

print("\n--- A VERIFICAR FICHEIROS ---")
try:
    if os.path.exists(CAMINHO_TABELA_KMS):
        DF_KMS = pd.read_excel(CAMINHO_TABELA_KMS, index_col=0)
        DF_EUROS = pd.read_excel(CAMINHO_TABELA_EUROS, index_col=0)
        print("-> Ficheiros de deslocações carregados (Formato Excel).")
    elif os.path.exists(caminho_csv_kms):
        DF_KMS = pd.read_csv(caminho_csv_kms, index_col=0)
        DF_EUROS = pd.read_csv(caminho_csv_euros, index_col=0)
        print("-> Ficheiros de deslocações carregados (Formato CSV).")
    else:
        print(f"-> AVISO CRÍTICO: Não encontrei as tabelas em {PASTA_DESTINO}!")
        
    if DF_KMS is not None:
        DF_KMS.index = DF_KMS.index.astype(str).str.strip()
        DF_KMS.columns = DF_KMS.columns.astype(str).str.strip()
        DF_EUROS.index = DF_EUROS.index.astype(str).str.strip()
        DF_EUROS.columns = DF_EUROS.columns.astype(str).str.strip()
except Exception as e:
    print(f"-> AVISO: Erro ao tentar ler as tabelas: {e}")
print("-----------------------------\n")

# ==========================================
# FUNCOES INTELIGENTES
# ==========================================
def definir_tipo_competicao(nome_prova):
    p = str(nome_prova).lower()
    if 'master' in p or 'distrital' in p or 'inter-associa' in p: 
        return 'Associação'
    if 'nacional' in p: 
        return 'Nacional'
    return 'Associação'

def extrair_genero(nome_prova):
    p = str(nome_prova).lower()
    if 'femin' in p or 'fem' in p: return 'Feminino'
    return 'Masculino' 

def extrair_fase(nome_prova):
    p = str(nome_prova).lower()
    if '2ª fase' in p or '2a fase' in p or 'ii fase' in p or 'final' in p: 
        return '2ª Fase'
    return '1ª Fase'

def extrair_escalao(nome_prova, tipo_jogo):
    p = str(nome_prova).lower()
    if 'master' in p: return 'Masters'
    if 'sub-18' in p or 'sub 18' in p or 'sub18' in p or 'u18' in p: return 'Sub18'
    if 'sub-16' in p or 'sub 16' in p or 'sub16' in p or 'u16' in p: return 'Sub16'
    if 'sub-14' in p or 'sub 14' in p or 'sub14' in p or 'u14' in p: return 'Sub14'
    
    if tipo_jogo == 'Nacional':
        if '2ª div' in p or '2a div' in p or 'ii div' in p or '2.ª div' in p: return 'CN2'
        if '1ª div' in p or '1a div' in p or 'i div' in p or 'senior' in p or 'sénior' in p: return 'CN1'
    else:
        if 'sub-21' in p or 'sub 21' in p or 'sub21' in p or 'u21' in p: return 'Sub21'
        if 'senior' in p or 'sénior' in p: return 'Seniores'
    return 'Outros'

def obter_preco_jogo(tipo_jogo, fase, genero, escalao):
    if MINHA_FUNCAO in TABELA_PREMIOS and tipo_jogo in TABELA_PREMIOS[MINHA_FUNCAO]:
        tabela_tipo = TABELA_PREMIOS[MINHA_FUNCAO][tipo_jogo]
        tabela_genero = tabela_tipo.get(genero, tabela_tipo.get("Masculino", {}))
        
        if tipo_jogo == "Associação":
            return tabela_genero.get(escalao, tabela_genero.get("Outros", 0.0))
        else:
            tabela_fase = tabela_genero.get(fase, tabela_genero.get("1ª Fase", {}))
            return tabela_fase.get(escalao, tabela_fase.get("Outros", 0.0))
    return 0.0

def obter_cidade(pavilhao):
    pav_limpo = str(pavilhao).lower()
    for palavra_chave, cidade in MAPA_PAVILHOES.items():
        if palavra_chave in pav_limpo:
            return cidade.strip()
    print(f"-> ATENÇÃO: O pavilhão '{pavilhao}' não está mapeado. Deslocação a zero.")
    return pavilhao.strip()

def obter_dados_deslocacao(origem, pavilhao):
    origem_limpa = origem.strip()
    cidade_destino = obter_cidade(pavilhao)
    
    if DF_KMS is None or DF_EUROS is None:
        return 0, 0.0, cidade_destino
        
    try:
        kms = float(DF_KMS.at[origem_limpa, cidade_destino])
        valor_cru = str(DF_EUROS.at[origem_limpa, cidade_destino]).replace('€', '').replace(',', '.').strip()
        return kms, float(valor_cru), cidade_destino
    except KeyError:
        print(f"-> ATENÇÃO: A base de dados não tem viagem de '{origem_limpa}' para '{cidade_destino}'.")
        return 0, 0.0, cidade_destino
    except Exception:
        return 0, 0.0, cidade_destino

# ==========================================
# O MOTOR DO SCRIPT E EXCEL
# ==========================================

def formatar_folha_excel(ws, df_origem, titulo_principal, is_nacional=False):
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_cabecalho = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    fill_azul_claro = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws['A1'] = titulo_principal
    ws['A1'].font = title_font
    ws.merge_cells('A1:K1')
    
    ws['A2'] = "MAPA DE PRESTAÇÃO DE CONTAS - PRESTAÇÃO DE SERVIÇOS E/OU DESLOCAÇÕES"
    ws['A2'].font = bold_font
    ws.merge_cells('A2:K2')

    ws['A4'] = f"Nome: {NOME_COMPLETO}"
    ws['A5'] = f"Localidade: {LOCAL_PARTIDA}"
    ws['A6'] = f"IBAN: {MEU_IBAN}"
    
    ws['F4'] = f"Categoria: {MINHA_FUNCAO}"
    ws['K4'] = f"Mês: {MES_PROCESSAMENTO}/{ANO_PROCESSAMENTO}"
    
    for row in range(4, 7):
        for col in ['A', 'F', 'K']:
            if ws[f'{col}{row}'].value:
                ws[f'{col}{row}'].font = bold_font

    larguras = {'A': 12, 'B': 8, 'C': 35, 'D': 12, 'E': 15, 'F': 25, 'G': 35, 'H': 8, 'I': 15, 'J': 12, 'K': 12}
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    max_row = ws.max_row
    
    if max_row >= 9:
        for row in ws.iter_rows(min_row=9, max_row=max_row, min_col=1, max_col=11):
            for cell in row:
                cell.border = thin_border
                
                if cell.row == 9:
                    cell.font = bold_font
                    cell.fill = fill_cabecalho
                    cell.alignment = Alignment(horizontal='center')
                elif cell.row > 9:
                    if cell.row % 2 == 0:
                        cell.fill = fill_azul_claro
                    if cell.column >= 8:
                        cell.alignment = Alignment(horizontal='center')

    # Calcula Totais
    total_deslocacoes = 0.0 if is_nacional else df_origem['Deslocacao Final'].sum() if not df_origem.empty else 0.0
    total_premios = df_origem['Premio Jogo'].sum() if not df_origem.empty else 0.0
    total_receber = total_premios + total_deslocacoes
    
    linha_total = max_row + 2
    
    ws[f'J{linha_total}'] = "TOTAL DESLOCAÇÕES:"
    ws[f'J{linha_total}'].font = bold_font
    ws[f'J{linha_total}'].alignment = Alignment(horizontal='right')
    ws[f'K{linha_total}'] = f"{total_deslocacoes:.2f} €"
    ws[f'K{linha_total}'].font = bold_font
    ws[f'K{linha_total}'].border = thin_border
    ws[f'K{linha_total}'].alignment = Alignment(horizontal='center')
    
    linha_total += 1
    ws[f'J{linha_total}'] = "TOTAL PRÉMIOS:"
    ws[f'J{linha_total}'].font = bold_font
    ws[f'J{linha_total}'].alignment = Alignment(horizontal='right')
    ws[f'K{linha_total}'] = f"{total_premios:.2f} €"
    ws[f'K{linha_total}'].font = bold_font
    ws[f'K{linha_total}'].border = thin_border
    ws[f'K{linha_total}'].alignment = Alignment(horizontal='center')

    linha_total += 1
    ws[f'J{linha_total}'] = "A RECEBER:"
    ws[f'J{linha_total}'].font = Font(bold=True, size=12)
    ws[f'J{linha_total}'].alignment = Alignment(horizontal='right')
    ws[f'K{linha_total}'] = f"{total_receber:.2f} €"
    ws[f'K{linha_total}'].font = Font(bold=True, size=12)
    ws[f'K{linha_total}'].border = thin_border
    ws[f'K{linha_total}'].alignment = Alignment(horizontal='center')
    ws[f'K{linha_total}'].fill = fill_cabecalho


def processar_tudo_automaticamente():
    print(f"A ligar ao site da FPB para o mes {MES_PROCESSAMENTO}/{ANO_PROCESSAMENTO}...")
    try:
        resposta = requests.get(URL_FPB, headers={'User-Agent': 'Mozilla/5.0'}, timeout=10)
        if resposta.status_code != 200: return print("Erro de acesso ao site.")
    except Exception as e:
        return print(f"Erro de conexão: {e}")

    try:
        tabelas = pd.read_html(StringIO(resposta.text))
        todos_os_jogos = pd.concat([tabelas[2], tabelas[3]], ignore_index=True)
    except Exception:
        return print("Erro: Nao encontrei as tabelas de jogos no site da FPB.")

    if isinstance(todos_os_jogos.columns, pd.MultiIndex):
        todos_os_jogos.columns = [c[-1] for c in todos_os_jogos.columns]
    todos_os_jogos.columns = [str(c).strip().replace('\xa0', ' ') for c in todos_os_jogos.columns]

    colunas = todos_os_jogos.columns

    col_prova = next((c for c in colunas if 'prova' in c.lower() or 'competi' in c.lower()), None)
    col_data = next((c for c in colunas if 'data' in c.lower()), None)
    col_pavilhao = next((c for c in colunas if 'pavilh' in c.lower() or 'local' in c.lower() or 'recinto' in c.lower()), None)
    col_equipas = next((c for c in colunas if 'equipa' in c.lower() or 'jogo' in c.lower()), None)

    if not all([col_prova, col_data, col_pavilhao]):
        return print("Erro: Nao consegui identificar as colunas base.")

    todos_os_jogos['Data_Python'] = pd.to_datetime(todos_os_jogos[col_data], dayfirst=True, errors='coerce')
    jogos_do_mes = todos_os_jogos[
        (todos_os_jogos['Data_Python'].dt.month == MES_PROCESSAMENTO) &
        (todos_os_jogos['Data_Python'].dt.year == ANO_PROCESSAMENTO)
    ].copy()

    if jogos_do_mes.empty:
        return print(f"Nenhum jogo encontrado para {MES_PROCESSAMENTO}/{ANO_PROCESSAMENTO}.")

    # ==========================================
    # CALCULOS INTERNOS E AGRUPAMENTO
    # ==========================================
    jogos_do_mes['Data_Apenas'] = jogos_do_mes['Data_Python'].dt.date
    jogos_do_mes['Hora_Limpa'] = jogos_do_mes['Data_Python'].dt.strftime('%H:%M')
    
    jogos_do_mes['Tipo'] = jogos_do_mes[col_prova].apply(definir_tipo_competicao)
    jogos_do_mes['Genero'] = jogos_do_mes[col_prova].apply(extrair_genero)
    jogos_do_mes['Fase'] = jogos_do_mes[col_prova].apply(extrair_fase)
    
    jogos_do_mes['Escalao'] = jogos_do_mes.apply(
        lambda l: extrair_escalao(l[col_prova], l['Tipo']), axis=1
    )
    
    jogos_do_mes['Premio Jogo'] = jogos_do_mes.apply(
        lambda l: obter_preco_jogo(l['Tipo'], l['Fase'], l['Genero'], l['Escalao']), axis=1
    )

    novas_colunas = jogos_do_mes.apply(
        lambda l: obter_dados_deslocacao(LOCAL_PARTIDA, l[col_pavilhao]), axis=1, result_type='expand'
    )
    jogos_do_mes['KMs'] = novas_colunas[0]
    jogos_do_mes['Valor Base'] = novas_colunas[1]
    jogos_do_mes['Cidade Destino'] = novas_colunas[2]

    # RESET AOS PAGADORES DE DESLOCACOES
    jogos_do_mes['Paga Deslocacao'] = False
    jogos_do_mes['Deslocacao Final'] = 0.0
    jogos_do_mes['Pagador Deslocacao'] = ''

    grupos_dias = jogos_do_mes.groupby(['Data_Apenas', 'Cidade Destino'])
    for (data, cidade), grupo in grupos_dias:
        tem_nacional = 'Nacional' in grupo['Tipo'].values
        
        if tem_nacional:
            idx_nacional = grupo[grupo['Tipo'] == 'Nacional'].index[0]
            for idx in grupo.index:
                if idx == idx_nacional:
                    jogos_do_mes.at[idx, 'Paga Deslocacao'] = True
                    jogos_do_mes.at[idx, 'Deslocacao Final'] = jogos_do_mes.at[idx, 'Valor Base']
                    jogos_do_mes.at[idx, 'Pagador Deslocacao'] = 'PAGADOR'
                else:
                    jogos_do_mes.at[idx, 'Paga Deslocacao'] = False
                    jogos_do_mes.at[idx, 'Deslocacao Final'] = 0.0
                    jogos_do_mes.at[idx, 'Pagador Deslocacao'] = 'PAGO FPB'
        else:
            idx_primeiro = grupo.index[0]
            for idx in grupo.index:
                if idx == idx_primeiro:
                    jogos_do_mes.at[idx, 'Paga Deslocacao'] = True
                    jogos_do_mes.at[idx, 'Deslocacao Final'] = jogos_do_mes.at[idx, 'Valor Base']
                    jogos_do_mes.at[idx, 'Pagador Deslocacao'] = 'PAGADOR'
                else:
                    jogos_do_mes.at[idx, 'Paga Deslocacao'] = False
                    jogos_do_mes.at[idx, 'Deslocacao Final'] = 0.0
                    jogos_do_mes.at[idx, 'Pagador Deslocacao'] = 'PAGO ABA'

    # ==========================================
    # SEPARAR OS JOGOS POR CATEGORIA E GERAR
    # ==========================================
    jogos_nacional = jogos_do_mes[jogos_do_mes['Tipo'] == 'Nacional'].copy()
    jogos_associacao = jogos_do_mes[(jogos_do_mes['Tipo'] == 'Associação') & (jogos_do_mes['Escalao'] != 'Masters')].copy()
    jogos_masters = jogos_do_mes[(jogos_do_mes['Tipo'] == 'Associação') & (jogos_do_mes['Escalao'] == 'Masters')].copy()

    def formata_deslocacao_nacional(row):
        if row['Paga Deslocacao'] and row['Deslocacao Final'] > 0:
            return f"{row['Deslocacao Final']:.2f} €"
        return ""

    def formata_deslocacao_associacao(row):
        if row['Paga Deslocacao'] and row['Deslocacao Final'] > 0:
            return f"{row['Deslocacao Final']:.2f} €"
        elif row['Pagador Deslocacao'] == 'PAGO FPB':
            return "PAGO FPB"
        elif row['Pagador Deslocacao'] == 'PAGO ABA':
            return "PAGO ABA"
        return ""

    def formata_total_linha(row, is_nacional):
        if is_nacional:
            return f"{row['Premio Jogo']:.2f} €"
        else:
            return f"{(row['Premio Jogo'] + row['Deslocacao Final']):.2f} €"

    def formatar_para_excel(df, is_nacional=False):
        return pd.DataFrame({
            "Data": df['Data_Python'].dt.strftime('%d/%m/%Y'),
            "Hora": df['Hora_Limpa'],
            "Prova": df[col_prova],
            "Escalão": df['Escalao'],
            "Função": MINHA_FUNCAO,
            "Recinto": df[col_pavilhao],
            "Jogo": df[col_equipas] if col_equipas else "",
            "Kms": df['KMs'].apply(lambda x: x if x > 0 else ""),
            "Deslocação": df.apply(lambda row: formata_deslocacao_nacional(row) if is_nacional else formata_deslocacao_associacao(row), axis=1),
            "Prémio (€)": df['Premio Jogo'].apply(lambda x: f"{x:.2f} €"),
            "Total (€)": df.apply(lambda row: formata_total_linha(row, is_nacional), axis=1)
        })

    caminho_base = f"{PASTA_DESTINO}\\Mapa_Contas_Geral_{MES_PROCESSAMENTO}_{ANO_PROCESSAMENTO}"
    caminho_final = f"{caminho_base}.xlsx"

    contador = 1
    while True:
        try:
            if os.path.exists(caminho_final):
                with open(caminho_final, 'a'): pass
            break
        except PermissionError:
            caminho_final = f"{caminho_base}_v{contador}.xlsx"
            contador += 1
    
    try:
        with pd.ExcelWriter(caminho_final, engine='openpyxl') as writer:
            escreveu_algo = False
            
            if not jogos_nacional.empty:
                df_nac = formatar_para_excel(jogos_nacional, is_nacional=True)
                df_nac.to_excel(writer, sheet_name="Nacional", index=False, startrow=8)
                escreveu_algo = True
            
            if not jogos_associacao.empty:
                df_assoc = formatar_para_excel(jogos_associacao, is_nacional=False)
                df_assoc.to_excel(writer, sheet_name="Associação", index=False, startrow=8)
                escreveu_algo = True
                
            if not jogos_masters.empty:
                df_mast = formatar_para_excel(jogos_masters, is_nacional=False)
                df_mast.to_excel(writer, sheet_name="Masters", index=False, startrow=8)
                escreveu_algo = True
                
            if not escreveu_algo:
                pd.DataFrame().to_excel(writer, sheet_name="Vazio")

        wb = openpyxl.load_workbook(caminho_final)
        
        if "Nacional" in wb.sheetnames:
            formatar_folha_excel(wb["Nacional"], jogos_nacional, "FEDERAÇÃO PORTUGUESA DE BASQUETEBOL", is_nacional=True)
            
        if "Associação" in wb.sheetnames:
            formatar_folha_excel(wb["Associação"], jogos_associacao, "ASSOCIAÇÃO DE BASQUETEBOL DE AVEIRO", is_nacional=False)
            
        if "Masters" in wb.sheetnames:
            formatar_folha_excel(wb["Masters"], jogos_masters, "ASSOCIAÇÃO DE AVEIRO - JOGOS MASTERS", is_nacional=False)

        try:
            if "Vazio" in wb.sheetnames:
                contagem = sum(1 for _ in wb.worksheets)
                if contagem > 1:
                    wb.remove(wb["Vazio"])
        except Exception:
            pass

        wb.save(caminho_final)
        
    except PermissionError:
        print(f"\n[ERRO CRÍTICO] O ficheiro Excel {caminho_final} está ABERTO!")
        print("Por favor, FECHA o Excel e volta a correr o script.")
        return

    def calc_total_limpo(df, is_nacional):
        if df.empty: return 0.0
        if is_nacional:
            return df['Premio Jogo'].sum()
        else:
            return df['Premio Jogo'].sum() + df['Deslocacao Final'].sum()

    print("\n" + "="*50)
    print(f"RELATORIO MENSAL CONCLUIDO ({MES_PROCESSAMENTO}/{ANO_PROCESSAMENTO})")
    print("="*50)
    if not jogos_nacional.empty:
        print(f"- Nacional: {calc_total_limpo(jogos_nacional, True):.2f} EUR (Viagens pagas extra pela FPB)")
    if not jogos_associacao.empty:
        print(f"- Associação: {calc_total_limpo(jogos_associacao, False):.2f} EUR")
    if not jogos_masters.empty:
        print(f"- Masters: {calc_total_limpo(jogos_masters, False):.2f} EUR")
    print("="*50)
    print(f"Ficheiro guardado em: {caminho_final}")

processar_tudo_automaticamente()